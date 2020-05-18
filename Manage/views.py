
import calendar
import os
from django.utils.translation import gettext as _

from openpyxl import Workbook,load_workbook
from openpyxl.styles import Color, Font,Border,Side,Alignment
from copy import copy
from django.shortcuts import render
from django.core.paginator import Paginator,PageNotAnInteger,EmptyPage

from django.http import JsonResponse, HttpResponseRedirect,HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.staticfiles.templatetags.staticfiles import static
from django.db.models import Q, Count, F, Min,Sum, Case,When ,Value,CharField
import operator
import functools

from django.utils import timezone, translation

from .forms import *
from app.models import *
from Receptionist.models import *
from Doctor.models import *
from Pharmacy.models import *

from dateutil import relativedelta


from django.views import View


# Create your views here.
def dash_board(request):

    dashboard_board = Board_Contents.objects.filter(use_yn='Y',board_type='BASIC',)[5:]


    return render(request,
        'Manage/dashboard.html',
            {
                'dashboard_board':dashboard_board,
            }
        )







@login_required
def manage(request):
    #doctor_search_form = DoctorsSearchForm()
    #
    #
    #list_exam_fee = []
    #list_precedures = []
    #list_radiologys = []
    #
    #
    #exam_fees = ExamFee.objects.filter(Q(code = 'E0010') | Q(code = 'E0011'))
    #for exam_fee in exam_fees:
    #    list_exam_fee.append({'code':exam_fee.code,'value':exam_fee.name})
    #
    #precedures = Precedure.objects.filter(code__contains='PM')
    #for precedure in precedures:
    #    list_precedures.append({'code':precedure.code,'value':precedure.name})
    #
    #radiologys = Precedure.objects.filter(code__contains='R', precedure_class_id = 10 )
    #for radiology in radiologys:
    #    list_radiologys.append({'code':radiology.code,'value':radiology.name})
    #
    #return render(request,
    #'Doctor/audit_PM.html',
    #    {
    #        'doctor_search':doctor_search_form,
    #
    #        'list_exam_fee':list_exam_fee,
    #        'list_precedures':list_precedures,
    #        'list_radiologys':list_radiologys,
    #
    #    }
    #)

    #####################################################

    #filters
    ##list_exam_fee = []
    ##list_lab = []
    ##list_precedure = []
    ##list_medicine = []
    ##
    ##
    ##    
    ##exams = ExamFee.objects.all().order_by('name')
    ##for exam in exams:
    ##    list_exam_fee.append({'code':exam.code,'value':exam.name})
    ##
    ##tests = Test.objects.all().order_by('name')
    ##for test in tests:
    ##    list_lab.append({'code':test.code,'value':test.name})
    ##
    ##precedures = Precedure.objects.all().order_by('name')
    ##for precedure in precedures:
    ##    list_precedure.append({'code':precedure.code,'value':precedure.name})
    ##    
    ##medicines = Medicine.objects.all().order_by('name')
    ##for medicine in medicines:
    ##    list_medicine.append({'code':medicine.code,'value':medicine.name})
    ##    



    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })
    #의사 정보 ? 
    doctor = Doctor.objects.values('name_short','id')

    #결제 방법
    payment_method = COMMCODE.objects.filter(use_yn = 'Y', commcode_grp='PAYMENT_METHOD',upper_commcode ='000006' ).annotate(code = F('commcode'),name = f_name ).values('code','name')

    #결제 상태
    payment_status = COMMCODE.objects.filter(use_yn = 'Y', commcode_grp='PAYMENT_STATUS',upper_commcode ='000006' ).annotate(code = F('commcode'),name = f_name ).values('code','name')

    return render(request,
        'Manage/manage.html',
            {
                'depart_medical':depart_medical,
                'doctor' : doctor,

                'payment_method':payment_method,
                'payment_status':payment_status,

                #'payment_search':payment_search_form,
                #'patient_search':patient_search_form,
                #'doctor_search':doctor_search_form,
                #'doctors':Doctor.objects.all(),
                #'medicine_search':medicine_search_form,

                #'list_exam_fee':list_exam_fee, # general will be precedure in template
                #'list_lab':list_lab,
                #'list_precedure':list_precedure,
                #'list_medicine':list_medicine,

            }
        )


def patient(request):
    return render(request,
        'Manage/patient.html',
            {

            }
        )

def payment(reqeust):
    return render(request,
        'Manage/payment.html',
            {

            }
        )



def search_payment(request):
    page_context = request.POST.get('page_context',10) # 페이지 컨텐츠 
    page = request.POST.get('page',1)

    date_start = request.POST.get('start')
    date_end = request.POST.get('end')

    depart = request.POST.get('depart')
    doctor = request.POST.get('doctor')

    payment_method = request.POST.get('payment_method')
    payment_status = request.POST.get('payment_status')

    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end, "%Y-%m-%d").date(), datetime.time.max)

    kwargs = {}

    kwargs['progress'] = 'done'
    if depart != '':
        kwargs['depart_id'] = depart
    if doctor != '':
        kwargs['doctor_id'] = doctor
    if payment_method !='':
        kwargs['payment__paymentrecord__method'] = payment_method
    if payment_status !='':
        kwargs['payment__progress'] = payment_status

    datas = []
    receptions = Reception.objects.filter(
            **kwargs ,
            recorded_date__range = (date_min, date_max), 
        ).prefetch_related(
            'diagnosis__exammanager_set',
            'diagnosis__testmanager_set',
            'diagnosis__preceduremanager_set',
            'diagnosis__medicinemanager_set',
            'payment__paymentrecord_set',
        ).order_by("-id")

    #print(receptions)
    payment_total = receptions.filter().aggregate(
        Sum('payment__sub_total'),
        Sum('payment__total'),
        Sum('payment__additional'),
        )

    


    #unpaid 구하기
    real_paid_total = receptions.aggregate(
        Sum('payment__paymentrecord__paid'),
        )
    payment_total_paid_amount =  0 if real_paid_total['payment__paymentrecord__paid__sum'] is None else real_paid_total['payment__paymentrecord__paid__sum']

    

    datas = []

    payment_total_subtotal = 0 if payment_total['payment__sub_total__sum'] is None else payment_total['payment__sub_total__sum']
    payment_total_additional = 0 if payment_total['payment__additional__sum'] is None else payment_total['payment__additional__sum']
    payment_total_total = 0 if payment_total['payment__total__sum'] is None else payment_total['payment__total__sum']
    payment_total_discount = payment_total_subtotal + payment_total_additional - payment_total_total
    payment_total_unpaid = payment_total_total - payment_total_paid_amount


    for reception in receptions:
        data = {
            'no':reception.id,
            'date':reception.recorded_date.strftime('%Y-%m-%d'),
            'Patient':reception.patient.name_kor,
            'patient_eng':reception.patient.name_eng,
            'date_of_birth':str(reception.patient.get_age()) + '/' + reception.patient.get_gender_simple(),
            'address':reception.patient.address,
            'gender':reception.patient.gender,
            'Depart':reception.depart.name,
            'Doctor_kor':reception.doctor.name_kor,
            'Doctor_eng': reception.doctor.name_eng,
            }

        list_exam_fee = []
        list_lab = []
        list_precedure= []
        list_radiation= []
        list_medicine= []




        #진료 아이템
        ##진료비
        tmp_exam_set = reception.diagnosis.exammanager_set.all()
        for tmp_exam in tmp_exam_set:
            list_exam_fee.append({
                'checked':tmp_exam.is_checked_discount,
                'code':tmp_exam.exam.code,
                'value':tmp_exam.exam.name
                })

        ##검사
        tmp_test_set = reception.diagnosis.testmanager_set.all()
        for tmp_test in tmp_test_set:
            list_lab.append({
                'checked':tmp_test.is_checked_discount,
                'code':tmp_test.test.code,
                'value':tmp_test.test.name,
                })


        ##처치 및 방사선
        tmp_precedure_set = reception.diagnosis.preceduremanager_set.all()
        for tmp_precedure in tmp_precedure_set:
            if 'R' in tmp_precedure.precedure.code:
                list_radiation.append({
                    'checked':tmp_precedure.is_checked_discount,
                    'code':tmp_precedure.precedure.code,
                    'value':tmp_precedure.precedure.name,
                    'amount':tmp_precedure.amount,
                    })
            else:
                list_precedure.append({
                    'checked':tmp_precedure.is_checked_discount,
                    'code':tmp_precedure.precedure.code,
                    'value':tmp_precedure.precedure.name,
                    'amount':tmp_precedure.amount,
                    })

        ##약
        tmp_medicine_set = reception.diagnosis.medicinemanager_set.all()
        for tmp_medicine in tmp_medicine_set:
            list_medicine.append({
                'checked':tmp_medicine.is_checked_discount,
                'code':tmp_medicine.medicine.code,
                'value':tmp_medicine.medicine.name,
                'amount':tmp_medicine.amount,
                })


        #수납 정보
        paid_by = '-'
        
        sub_total = reception.payment.sub_total
        additional = 0 if reception.payment.additional is None else reception.payment.additional
        discount = reception.payment.discounted_amount
        total = reception.payment.total
        unpaid = 0

        if discount is None:
            discount_percent = reception.payment.discounted
            if discount_percent is None:
                discount = 0
            else:
                discount = discount_percent / 100 * sub_total
        
        #수납 상태 및 방법
        #if reception.payment.progress != 'paid':
        list_paid_record = reception.payment.paymentrecord_set.all()
        all_paid = 0
            
        paid_by = ''
        paid_by_remit = False
        paid_by_card = False
        paid_by_cash = False

        #print(reception.payment.total)



        #if list_paid_record.count() != 0:
        for paid_record in list_paid_record:
            all_paid += paid_record.paid
            if paid_record.method == 'cash':
                paid_by_cash = True
            if paid_record.method == 'card':
                paid_by_card = True
            if paid_record.method == 'remit':
                paid_by_remit = True
        unpaid = total - all_paid
        #else:
        #    unpaid = total
                
            
        
            

        if paid_by_cash is True:
            paid_by += 'Cash<br/>'
        if paid_by_card is True:
            paid_by += 'Card<br/>'
        if paid_by_remit is True:
            paid_by += 'Remit<br/>'




        data.update({
            'list_exam_fee':list_exam_fee,
            'list_lab':list_lab,
            'list_precedure':list_precedure,
            'list_radiation':list_radiation,
            'list_medicine':list_medicine,

            'paid_by':paid_by,
            'sub_total':sub_total,
            'additional':additional,
            'discount':discount,
            'total':total,
            'unpaid':unpaid,

            })

        datas.append(data)

    paginator = Paginator(datas, page_context)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)


    context = {
            'datas':list(paging_data),
            'page_range_start':paging_data.paginator.page_range.start,
            'page_range_stop':paging_data.paginator.page_range.stop,
            'page_number':paging_data.number,
            'has_previous':paging_data.has_previous(),
            'has_next':paging_data.has_next(),

 

            'payment_total_subtotal':payment_total_subtotal,
            'payment_total_additional':payment_total_additional,
            'payment_total_discount':payment_total_discount,
            'payment_total_total':payment_total_total,
            'payment_total_unpaid':payment_total_unpaid,
            }
    
    return JsonResponse(context)

    #filter_general = request.POST.get('general')
    #filter_medicine = request.POST.get('medicine')employee_add_edit_get
    #filter_lab = request.POST.get('lab')
    #
    #pup =request.POST.get('pup')
    #paid_by = request.POST.get('paid_by')
    #
    #date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    #date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end, "%Y-%m-%d").date(), datetime.time.max)
    #
    #
    #
    #
    #
    #if depart != '':
    #    kwargs.update({'depart_id':depart})
    #if doctor != '':
    #    kwargs.update({'doctor_id':doctor})
    #
    #
    #datas = []
    #receptions = Reception.objects.filter(**kwargs ,recorded_date__range = (date_min, date_max), progress = 'done').order_by("-id")
    #
    #
    #page = request.POST.get('page',1)
    #payment_total_total = 0
    #payment_total_paid = 0
    #payment_total_unpaid = 0
    #for reception in receptions:
    #    data = {}
    #    try:
    #        general = []
    #        lab = []
    #        medi = []
    #        scaling = []
    #        panorama = []
    #        total_payment=0
    #        if filter_general == '' and filter_medicine == '' and filter_lab == '':
    #            tmp_exam_set = reception.diagnosis.exammanager_set.all()
    #            for tmp_exam in tmp_exam_set:
    #                if hasattr(tmp_exam,'doctor'):
    #                    general.append({
    #                        'code':tmp_exam.exam.code,
    #                        'value':tmp_exam.name + tmp_exam.exam.doctor.name_kor
    #                        })
    #                else:
    #                    general.append({
    #                        'code':tmp_exam.exam.code,
    #                        'value':tmp_exam.exam.name
    #                        })
    #       
    #            tmp_test_set = reception.diagnosis.testmanager_set.all()
    #            for tmp_test in tmp_test_set:
    #                lab.append({
    #                    'code':tmp_test.test.code,
    #                    'value':tmp_test.test.name,
    #                    })
    #
    #            tmp_precedure_set = reception.diagnosis.preceduremanager_set.all()
    #            for tmp_precedure in tmp_precedure_set:
    #                if 'scaling' in tmp_precedure.precedure.name.lower():
    #                    scaling.append({
    #                        'code':tmp_precedure.precedure.code,
    #                        'value':tmp_precedure.precedure.name
    #                        })
    #
    #                elif 'injection' in tmp_precedure.precedure.name.lower():
    #                    general.append({
    #                        'code':tmp_precedure.precedure.code,
    #                        'value':tmp_precedure.precedure.name
    #                        })
    #
    #                elif 'panorama' in tmp_precedure.precedure.name.lower():
    #                    panorama.append({
    #                        'code':tmp_precedure.precedure.code,
    #                        'value':tmp_precedure.precedure.name
    #                        })
    #
    #                else:
    #                    general.append({
    #                        'code':tmp_precedure.precedure.code,
    #                        'value':tmp_precedure.precedure.name
    #                        })
    #        
    #
    #            tmp_medicine_set = reception.diagnosis.medicinemanager_set.all()
    #            for tmp_medicine in tmp_medicine_set:
    #                if tmp_medicine.medicine.medicine_class_id is 31:
    #                    general.append({
    #                        'code':tmp_medicine.medicine.code,
    #                        'value':tmp_medicine.medicine.name + ' x ' + str(tmp_medicine.days * tmp_medicine.amount),
    #                        })
    #                else:
    #                    medi.append({
    #                        'code':tmp_medicine.medicine.code,
    #                        'value':tmp_medicine.medicine.name + ' x ' + str(tmp_medicine.days * tmp_medicine.amount),
    #                        })
    #
    #
    #        else:
    #            if filter_general != '':
    #                if 'E' in filter_general:
    #                    tmp_exam_set = reception.diagnosis.exammanager_set.all()
    #                    res = True
    #                    for tmp_exam_data in tmp_exam_set:
    #                        if 'E_NEW' in filter_general:
    #                            if 'New' not in tmp_exam_data.exam.name:
    #                                res = False
    #                        elif 'E_REP' in filter_general:
    #                            if 'Rep' not in tmp_exam_data.exam.name:
    #                                res = False
    #                        elif 'E_DNT' in filter_general:
    #                            if 'Ora' not in tmp_exam_data.exam.name:
    #                                res = False
    #                        else:
    #                            tmp_exam = ExamFee.objects.get(code = filter_general)
    #                            if tmp_exam.code not in tmp_exam_data.exam.code:
    #                                res = False
    #                        
    #                        if res:
    #                            general.append({
    #                                'code':tmp_exam_data.exam.code,
    #                                'value':tmp_exam_data.exam.name
    #                                })
    #                            total_payment += tmp_exam_data.exam.get_price(reception.recorded_date)
    #                    if res is False or tmp_exam_set.count()==0:
    #                        continue
    #                            
    #
    #                elif 'MR' in filter_general:
    #                    tmp_exam = ExamFee.objects.get(code = filter_general)
    #                    tmp_exam_set = reception.diagnosis.exammanager_set.filter(exam_id = tmp_exam.id)
    #                    if tmp_exam_set.count() == 0:
    #                        continue
    #                    for tmp_exam_data in tmp_exam_set:
    #                        general.append({
    #                            'code':tmp_exam_data.exam.code,
    #                            'value':tmp_exam_data.exam.name
    #                            })
    #                        total_payment += tmp_exam_data.exam.get_price(reception.recorded_date)
    #
    #                elif 'M' in filter_general:
    #                    tmp_medicine = Medicine.objects.get(code = filter_general)
    #                    tmp_medi_set = reception.diagnosis.medicinemanager_set.filter(medicine_id = tmp_medicine.id)
    #
    #                    if tmp_medi_set.count() == 0:
    #                        continue
    #                    for tmp_medi_data in tmp_medi_set:
    #                        general.append({
    #                            'code':tmp_medi_data.medicine.code,
    #                            'value':tmp_medi_data.medicine.name + ' x ' + str(tmp_medi_data.days * tmp_medi_data.amount),
    #                            })
    #                        total_payment += tmp_medi_data.medicine.get_price(reception.recorded_date) * tmp_medi_data.days * tmp_medi_data.amount
    #
    #                else: #P D G R U O OB
    #                    tmp_precedure = Precedure.objects.get(code = filter_general)
    #                    tmp_precedure_set = reception.diagnosis.preceduremanager_set.filter(precedure_id = tmp_precedure.id)
    #
    #                    if tmp_precedure_set.count() == 0:
    #                        continue
    #                    for tmp_precedure_data in tmp_precedure_set:
    #                        general.append({
    #                                'code':tmp_precedure_data.precedure.code,
    #                                'value':tmp_precedure_data.precedure.name
    #                                })
    #                        total_payment += tmp_precedure_data.precedure.get_price(reception.recorded_date)
    #
    #            if filter_medicine != '':
    #                tmp_medicine = Medicine.objects.get(code = filter_medicine)
    #                tmp_set = reception.diagnosis.medicinemanager_set.filter(medicine_id = tmp_medicine.id)
    #
    #                if tmp_set.count() == 0:
    #                        continue
    #                for tmp_data in tmp_set:
    #                    medi.append({
    #                        'code':tmp_data.exam.code,
    #                        'value':tmp_data.medicine.name + ' x ' + str(tmp_data.days * tmp_data.amount),
    #                        })
    #                    total_payment += tmp_precedure_data.precedure.get_price(reception.recorded_date)
    #            if filter_lab != '':
    #                tmp_test = Test.objects.get(code = filter_lab)
    #                tmp_set = reception.diagnosis.testmanager_set.filter(test_id = tmp_test.id)
    #                if tmp_set.count() == 0:
    #                        continue
    #                for tmp_data in tmp_set:
    #                    lab.append({
    #                       'code':tmp_data.exam.code,
    #                        'value':tmp_data.exam.name
    #                        })
    #                    total_payment += tmp_precedure_data.precedure.get_price(reception.recorded_date)
    #            
    #
    #        data.update({'general':general})
    #        data.update({'medi':medi})
    #        data.update({'lab':lab})
    #        data.update({'scaling':scaling})
    #        data.update({'panorama':panorama})
    #
    #        paid_set = reception.payment.paymentrecord_set.all()
    #        paid_sum = 0
    #        
    #        for paid in paid_set:
    #            paid_sum += paid.paid
    #
    #        unpaid_sum = reception.payment.total - paid_sum
    #    
    #        if pup == 'Paid':
    #            if unpaid_sum != 0:
    #                continue
    #        elif pup == 'Unpaid':
    #            if unpaid_sum == 0:
    #                continue
    #
    #        if filter_general == '' and filter_medicine == '' and filter_lab == '':
    #            payment_total_total += reception.payment.total
    #            payment_total_paid += reception.payment.total - unpaid_sum
    #            payment_total_unpaid += unpaid_sum
    #            total_payment = reception.payment.total
    #        else:
    #            payment_total_total += total_payment
    #            paid_sum = 0
    #            unpaid_sum = 0
    #
    #        data.update({
    #            'no':reception.id,
    #            'date':reception.recorded_date.strftime('%d-%b-%y'),
    #            'Patient':reception.patient.name_kor,
    #            'patient_eng':reception.patient.name_eng,
    #            'date_of_birth':str(reception.patient.get_age()) + '/' + reception.patient.get_gender_simple(),
    #            'address':reception.patient.address,
    #            'gender':reception.patient.gender,
    #            'Depart':reception.depart.name,
    #            'Doctor':reception.doctor.get_name(),
    #
    #            'paid_by_cash':'',
    #            'paid_by_card':'',
    #            'paid_by_remit':'',
    #
    #            'total' :total_payment,
    #            'paid':paid_sum,
    #            'unpaid':unpaid_sum,
    #            })
    #
    #        
    #
    #
    #        pay_records = PaymentRecord.objects.filter(payment = reception.payment)
    #
    #        for pay_record in pay_records:
    #            if pay_record.method == 'card':
    #                data.update({'paid_by_card':'card'})
    #            elif pay_record.method == 'cash':
    #                data.update({'paid_by_card':'cash'})
    #            elif pay_record.method == 'remit':
    #                data.update({'paid_by_card':'remit'})
    #
    #        datas.append(data)
    #    except Diagnosis.DoesNotExist:
    #        pass
    #
    #paginator = Paginator(datas, page_context)
    #try:
    #    paging_data = paginator.page(page)
    #except PageNotAnInteger:
    #    paging_data = paginator.page(1)
    #except EmptyPage:
    #    paging_data = paginator.page(paginator.num_pages)
    #
    #
    #context = {
    #           'datas':list(paging_data),
    #           'page_range_start':paging_data.paginator.page_range.start,
    #           'page_range_stop':paging_data.paginator.page_range.stop,
    #           'page_number':paging_data.number,
    #           'has_previous':paging_data.has_previous(),
    #           'has_next':paging_data.has_next(),
    #
    #           #for graph
    #           'days':(date_max - date_min).days +1 ,
    #
    #           'payment_total_total':payment_total_total,
    #           'payment_total_paid':payment_total_paid,
    #           'payment_total_unpaid':payment_total_unpaid,
    #
    #           }
    #
    #return JsonResponse(context)

@login_required
def doctor_profit(request):

    page_context = request.POST.get('page_context',10) # 페이지 컨텐츠 
    kwargs = {}
    datas = []

    doctor = request.POST.get('doctor')

    date_start = request.POST.get('start_date')
    date_end = request.POST.get('end_date')


    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end, "%Y-%m-%d").date(), datetime.time.max)
    
    page = request.POST.get('page',1)
    
    if doctor != '':
        kwargs.update({'doctor_id':doctor})

    context = {}

    if request.user.is_admin or request.user.doctor.depart.name == 'PM' :
        total_amount = 0
        total_additional = 0
        amount_discount = 0

        filter_search = request.POST.get('search')

        receptions = Reception.objects.filter(**kwargs ,recorded_date__range = (date_min, date_max), progress = 'done').select_related('payment').filter(payment__progress='paid').order_by("-id")
        #receptions = Reception.objects.filter(**kwargs ,recorded_date__range = (date_min, date_max), progress = 'done').select_related('diagnosis').select_related('payment').order_by("-id")
        for reception in receptions:
            exams = []
            precedures = []
            radiographys = []

            data={}
            sub_total = 0


            if filter_search == '':
                tmp_exam_set = reception.diagnosis.exammanager_set.all()
                tmp_precedure_set = reception.diagnosis.preceduremanager_set.all()
            else:
                if 'E' in filter_search:
                    tmp_exam_set = reception.diagnosis.exammanager_set.prefetch_related('exam').filter(exam__code=filter_search)
                    tmp_precedure_set = reception.diagnosis.preceduremanager_set.none()
                elif 'PM' in filter_search:
                    tmp_exam_set = reception.diagnosis.exammanager_set.none()
                    tmp_precedure_set = reception.diagnosis.preceduremanager_set.all().prefetch_related('precedure').filter(precedure__code=filter_search)

                elif 'R' in filter_search:
                    tmp_exam_set = reception.diagnosis.exammanager_set.none()
                    tmp_precedure_set = reception.diagnosis.preceduremanager_set.all().prefetch_related('precedure').filter(precedure__code__icontains='R')

                if tmp_exam_set.count() is 0 and tmp_precedure_set.count() is 0:
                    continue
                
            for tmp_exam in tmp_exam_set:
                    exams.append({
                        'code':tmp_exam.exam.code,
                        'value':tmp_exam.exam.name,
                        })
                    sub_total += tmp_exam.exam.get_price(reception.recorded_date)
                    total_amount += sub_total
            for precedure_set in tmp_precedure_set:
                    if 'R' in precedure_set.precedure.code:
                        radiographys.append({
                            'code':precedure_set.precedure.code,
                            'value':precedure_set.precedure.name,
                            'amount':precedure_set.amount,
                            })
                        sub_total += precedure_set.precedure.get_price(reception.recorded_date)
                        total_amount += sub_total
                    else:
                        precedures.append({
                            'code':precedure_set.precedure.code,
                            'value':precedure_set.precedure.name,
                            })
                        sub_total += precedure_set.precedure.get_price(reception.recorded_date)
                        total_amount += sub_total

            total_additional += 0 if reception.payment.additional is None else reception.payment.additional
            
            if reception.payment.discounted is None:
                discount = reception.payment.discounted_amount if filter_search == '' else 0
            else:
                discount = reception.payment.discounted / 100 * reception.payment.sub_total
            amount_discount += 0 if discount is None else discount

            data.update({
                'exams':exams,
                'precedures':precedures,
                'radiographys':radiographys,
                
                'subtotal':reception.payment.sub_total if filter_search == '' else sub_total,
                'discount':discount if filter_search == '' else 0,
                'total':reception.payment.total if filter_search == '' else sub_total,
                'additional':reception.payment.additional if filter_search == '' else 0,

  
                'no':reception.id,
                'date':reception.recorded_date.strftime('%d-%b-%y'),
                'Patient':reception.patient.name_kor,
                'patient_eng':reception.patient.name_eng,
                'date_of_birth':str(reception.patient.get_age()) + '/' + reception.patient.get_gender_simple(),
                'address':reception.patient.address,
                'gender':reception.patient.gender,
                'Depart':reception.depart.name,
                'Doctor':reception.doctor.get_name(),
                })

            
            datas.append(data)

        context.update({
            'total_amount':total_amount,
            'total_additional':total_additional,
            })
    else:
        filter_exam_fee = request.POST.get('exam_fee')
        filter_test = request.POST.get('test')
        filter_precedure = request.POST.get('precedure')
        filter_medicine = request.POST.get('medicine')

        pup =request.POST.get('pup')
        paid_by = request.POST.get('paid_by')


        receptions = Reception.objects.filter(**kwargs ,recorded_date__range = (date_min, date_max), progress = 'done').select_related('payment').filter(payment__progress='paid').order_by("-id")

        
        
        amount_exam_fee = 0
        amount_test = 0
        amount_precedure = 0
        amount_medicine = 0

        
        
        for reception in receptions:
            data = {}
            try:
                exam_fee = []
                test = []
                precedure = []
                medi = []

                
                #필터링 없을 때
                if filter_exam_fee == '' and filter_test == '' and filter_precedure == '' and filter_medicine=='':
                   
                    tmp_exam_set = reception.diagnosis.exammanager_set.all()
                    for tmp_exam in tmp_exam_set:
                        exam_fee.append({
                            'code':tmp_exam.exam.code,
                            'value':tmp_exam.exam.name
                            })

                    tmp_test_set = reception.diagnosis.testmanager_set.all()
                    for tmp_test in tmp_test_set:
                        test.append({
                            'code':tmp_test.test.code,
                            'value':tmp_test.test.name,
                            })

                    tmp_precedure_set = reception.diagnosis.preceduremanager_set.all()
                    for tmp_precedure in tmp_precedure_set:
                        precedure.append({
                            'code':tmp_precedure.precedure.code,
                            'value':tmp_precedure.precedure.name
                            })
            

                    tmp_medicine_set = reception.diagnosis.medicinemanager_set.all()
                    for tmp_medicine in tmp_medicine_set:
                        medi.append({
                            'code':tmp_medicine.medicine.code,
                            'value':tmp_medicine.medicine.name + ' x ' + str(tmp_medicine.days * tmp_medicine.amount),
                            })

                paid_set = reception.payment.paymentrecord_set.all()
                paid_sum = 0
                for paid in paid_set:
                    paid_sum  += paid.paid

                data.update({
                    'no':reception.id,
                    'date':reception.recorded_date.strftime('%d-%m-%Y'),
                    'Patient':reception.patient.name_kor,
                    'patient_eng':reception.patient.name_eng,
                    'date_of_birth':str(reception.patient.get_age()) + '/' + reception.patient.get_gender_simple(),
                    'address':reception.patient.address,
                    'gender':reception.patient.gender,
                    'Depart':reception.depart.name,
                    'Doctor':reception.doctor.get_name(),

                    'list_exam_fee':exam_fee,
                    'list_test':test,
                    'list_precedure':precedure,
                    'list_medi':medi,

                    'paid_by_cash':'',
                    'paid_by_card':'',
                    'paid_by_remit':'',

                    'sub_total':reception.payment.sub_total,
                    'total' :reception.payment.total,
                    'discount':reception.payment.discounted_amount,
                    })


                for pay_record in paid_set:
                    if pay_record.method == 'card':
                        data.update({'paid_by_card':'card'})
                    elif pay_record.method == 'cash':
                        data.update({'paid_by_card':'cash'})
                    elif pay_record.method == 'remit':
                        data.update({'paid_by_card':'remit'})

                datas.append(data)
            except Diagnosis.DoesNotExist:
                pass


        if (date_max - date_min).days == 0:  #단일 날짜는 당일을 Subtotal / 선택된 달의 금액을 Total
             first_day = datetime.datetime.strptime(date_start, "%Y-%m-%d").date().replace(day=1)
             last_day = (first_day + relativedelta.relativedelta(months=1)) - datetime.timedelta(seconds=1)
             
             receptions = Reception.objects.filter(**kwargs ,recorded_date__range = (first_day, last_day), progress = 'done').order_by("-id").select_related('payment')
             monthly_total = 0
             for reception in receptions:
                 monthly_total += reception.payment.total
             context.update({
                'monthly_total':monthly_total,
                })
        #else: #범위 날짜는 SubTotal 무시 /범위 계산을 Total로

    query_total = Reception.objects.filter(**kwargs ,recorded_date__range = (date_min, date_max), progress = 'done').select_related('payment').filter(payment__progress='paid')


    
    query_total = receptions.aggregate(
        amount_sub_total=Sum('payment__sub_total'),
        #amount_discount = Sum('payment__discounted_amount'),
        amount_total = Sum('payment__total'),
        )

    context.update({
        'amount_sub_total':query_total['amount_sub_total'] if query_total['amount_sub_total'] is not None else 0,
        'amount_discount':amount_discount,
        'amount_total':query_total['amount_total'] if query_total['amount_total'] is not None else 0,
        })
    
    paginator = Paginator(datas, page_context)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)
    

    context.update({
                'datas':list(paging_data),
                'page_range_start':paging_data.paginator.page_range.start,
                'page_range_stop':paging_data.paginator.page_range.stop,
                'page_number':paging_data.number,
                'has_previous':paging_data.has_previous(),
                'has_next':paging_data.has_next(),

                #for graph
                'days':(date_max - date_min).days +1 ,


                })
    return JsonResponse(context)



def audit_excel(request):


    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')

    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end, "%Y-%m-%d").date(), datetime.time.max)

    kwargs = {}
    kwargs['progress'] = 'done'

    datas = []
    receptions = Reception.objects.filter(
            **kwargs ,
            recorded_date__range = (date_min, date_max), 
        ).prefetch_related(
            'diagnosis__exammanager_set',
            'diagnosis__testmanager_set',
            'diagnosis__preceduremanager_set',
            'diagnosis__medicinemanager_set',
            'payment__paymentrecord_set',
        ).order_by("-id")

    #이름 설정
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="AUDIT REPORT ' + date_start + '_' + date_end +'.xlsx"'
    
    #기본 양식 경로
    path = static('excel_form/I-MEDICARE_REPORT.xlsx')

    #엑셀 파일 불러오기
    wb = load_workbook('static/excel_form/audit_report.xlsx') #Workbook()
    ws = wb.active# grab the active worksheet

    #선택한 날짜
    ws['B3'] = date_start + ' - ' + date_end

    #처음 시작 A7 ~ W7
    current_row = 7
    data_num = 1
    for reception in receptions:
        recorded_date = reception.recorded_date
        #기본 정보
        ws['A' + str(current_row)] = data_num
        ws['B' + str(current_row)] = reception.recorded_date.strftime('%Y-%m-%d')
        ws['C' + str(current_row)] = reception.patient.get_chart_no()
        ws['D' + str(current_row)] = reception.patient.name_kor + ' / ' + reception.patient.name_eng
        ws['E' + str(current_row)] = reception.depart.name
        ws['F' + str(current_row)] = reception.doctor.name_short



        #진료 아이템
        ##진료비
        temp_row = current_row
        highest = 1
        tmp_exam_set = reception.diagnosis.exammanager_set.all()
        for tmp_exam in tmp_exam_set:
            ws['G' + str(temp_row)] = tmp_exam.exam.name
            ws['H' + str(temp_row)] = "-"
            ws['I' + str(temp_row)] = tmp_exam.exam.get_price(recorded_date)

            temp_row += 1
  
        if tmp_exam_set.count() > highest:
            highest = tmp_exam_set.count() 

        ##약
        temp_row = current_row
        tmp_medicine_set = reception.diagnosis.medicinemanager_set.all()
        for tmp_medicine in tmp_medicine_set:
            ws['J' + str(temp_row)] = tmp_medicine.medicine.name
            ws['K' + str(temp_row)] = tmp_medicine.amount
            ws['L' + str(temp_row)] = tmp_medicine.amount * tmp_medicine.medicine.get_price(recorded_date)

            temp_row += 1

        if tmp_medicine_set.count() > highest:
            highest = tmp_medicine_set.count() 


        ##검사
        temp_row = current_row
        tmp_test_set = reception.diagnosis.testmanager_set.all()
        for tmp_test in tmp_test_set:
            ws['M' + str(temp_row)] = tmp_test.test.name
            ws['N' + str(temp_row)] = "-"
            ws['O' + str(temp_row)] = tmp_test.test.get_price(recorded_date)

            temp_row += 1
        
        if tmp_test_set.count() > highest:
            highest = tmp_test_set.count() 


        ##처치 및 방사선
        tmp_precedure_set = reception.diagnosis.preceduremanager_set.all()
        temp_row = current_row
        temp_row_p = current_row
        for tmp_precedure in tmp_precedure_set:
            if 'R' in tmp_precedure.precedure.code:
                ws['P' + str(temp_row)] = tmp_precedure.precedure.name
                ws['Q' + str(temp_row)] = tmp_precedure.amount
                ws['R' + str(temp_row)] = tmp_precedure.precedure.get_price(recorded_date)

                temp_row += 1
        
            else:
                ws['S' + str(temp_row_p)] = tmp_precedure.precedure.name
                ws['T' + str(temp_row_p)] = tmp_precedure.amount
                ws['U' + str(temp_row_p)] = tmp_precedure.precedure.get_price(recorded_date)

                temp_row_p += 1
        
        if tmp_precedure_set.filter(precedure__code__icontains='R').count() > highest:
            highest = tmp_precedure_set.filter(precedure__code__icontains='R').count()
        
        if tmp_precedure_set.exclude(precedure__code__icontains='R').count() > highest:
            highest = tmp_precedure_set.exclude(precedure__code__icontains='R').count()
        
        
        paid_sum = reception.payment.paymentrecord_set.aggregate(Sum('paid')).get('paid__sum')

        ws['V' + str(current_row)] = reception.payment.sub_total
        if reception.payment.discounted != 0:
            ws['W' + str(current_row)] = reception.payment.sub_total / 100 * reception.payment.discounted
        else:
            ws['W' + str(current_row)] = reception.payment.discounted_amount
        ws['X' + str(current_row)] = reception.payment.additional
        ws['Y' + str(current_row)] = reception.payment.total
        ws['Z' + str(current_row)] = reception.payment.total - paid_sum
        ws['AA' + str(current_row)] = paid_sum
        ws['AB' + str(current_row)] = reception.payment.memo 


        if highest != 0:
            ws.merge_cells('A' + str(current_row) + ':A' + str(current_row + highest-1))
            ws.merge_cells('B' + str(current_row) + ':B' + str(current_row + highest-1))
            ws.merge_cells('C' + str(current_row) + ':C' + str(current_row + highest-1))
            ws.merge_cells('D' + str(current_row) + ':D' + str(current_row + highest-1))
            ws.merge_cells('E' + str(current_row) + ':E' + str(current_row + highest-1))
            ws.merge_cells('F' + str(current_row) + ':F' + str(current_row + highest-1))
            ws.merge_cells('V' + str(current_row) + ':V' + str(current_row + highest-1))
            ws.merge_cells('W' + str(current_row) + ':W' + str(current_row + highest-1))
            ws.merge_cells('X' + str(current_row) + ':X' + str(current_row + highest-1))
            ws.merge_cells('Y' + str(current_row) + ':Y' + str(current_row + highest-1))
            ws.merge_cells('Z' + str(current_row) + ':Z' + str(current_row + highest-1))
            ws.merge_cells('AA' + str(current_row) + ':AA' + str(current_row + highest-1))
            ws.merge_cells('AB' + str(current_row) + ':AB' + str(current_row + highest-1))

        current_row += highest 
        data_num +=1

    border_thin = Border(top=Side(border_style="thin", color="000000") ,
                        left=Side(border_style="thin", color="000000") ,
                       right=Side(border_style="thin", color="000000") ,
                      bottom=Side(border_style="thin", color="000000") )

    rows = ws['A7:AB' + str(current_row)]
    for row in rows:
        for cell in row:
            cell.border = border_thin



    wb.save(response)
    return response


def rec_report_excel(request):
    
    date_start = datetime.datetime.today().strftime("%Y-%m-%d")

    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.max)

    kwargs = {}
    kwargs['progress'] = 'done'

    datas = []
    receptions = Reception.objects.filter(
            **kwargs ,
            recorded_date__range = (date_min, date_max), 
        ).prefetch_related(
            'diagnosis__exammanager_set',
            'diagnosis__testmanager_set',
            'diagnosis__preceduremanager_set',
            'diagnosis__medicinemanager_set',
            'payment__paymentrecord_set',
        ).order_by("-id")

    #이름 설정
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="RECEPTION REPORT ' + date_start +'.xlsx"'

    #엑셀 파일 불러오기
    wb = load_workbook('static/excel_form/reception_report.xlsx') #Workbook()
    ws = wb.active# grab the active worksheet

    #선택한 날짜
    ws['B1'] = 'BÁO CÁO DOANH THU PHÒNG KHÁM NGÀY ' + date_min.strftime('%d.%m.%Y')


    #처음 시작 A7 ~ W7
    current_row = 7
    form_row = 8
    data_num = 1

    #스타일 초기화
    style_font = ws['A7'].font
    style_border = ws['A7'].border
    style_fill = ws['A7'].fill
    #string
    style_string_format = ws['A7'].number_format
    #number
    style_number_format = ws['G7'].number_format
    style_protection = ws['A7'].protection
    style_alignment = ws['A7'].alignment


    #값 초기화
    total_subtotal = 0
    total_discount = 0
    total_total = 0
    total_paid_cash = 0
    total_paid_transfer = 0 
    total_debit = 0


    for reception in receptions:
        recorded_date = reception.recorded_date
        ws.insert_rows(current_row+1)
        
        rows = ws['A' + str(current_row) + ':T' + str(current_row)]
        new_rows =ws['A' + str(current_row + 1) + ':T' + str(current_row + 1)]
        ws['A7'].style# = rows.style

        for row in rows:
            for cell in row:
                ws.cell(row = cell.row+1, column=cell.column).font = copy(cell.font)
                ws.cell(row = cell.row+1, column=cell.column).border = copy(cell.border)
                ws.cell(row = cell.row+1, column=cell.column).fill = copy(cell.fill)
                ws.cell(row = cell.row+1, column=cell.column).number_format = copy(cell.number_format)
                ws.cell(row = cell.row+1, column=cell.column).protection = copy(cell.protection)
                ws.cell(row = cell.row+1, column=cell.column).alignment = copy(cell.alignment)

        #기본 정보
        ws['A' + str(current_row)] = data_num
        ws['B' + str(current_row)] = reception.recorded_date.strftime('%Y-%m-%d')
        ws['C' + str(current_row)] = reception.depart.name
        ws['D' + str(current_row)] = reception.patient.name_kor + ' / ' + reception.patient.name_eng


        ws['E' + str(current_row)] = reception.diagnosis.diagnosis

        ws['F' + str(current_row)] = reception.payment.sub_total
        total_subtotal += reception.payment.sub_total
        if reception.payment.discounted != 0:
            ws['G' + str(current_row)] = reception.payment.sub_total / 100 * reception.payment.discounted
            total_discount += reception.payment.sub_total / 100 * reception.payment.discounted
        else:
            ws['G' + str(current_row)] = reception.payment.discounted_amount
            total_discount += reception.payment.discounted_amount
        ws['H' + str(current_row)] = reception.payment.total
        total_total += reception.payment.total

        ws['I' + str(current_row)] = ''

        #실제 지불금
        if reception.payment.paymentrecord_set.count() != 0:
            paid = reception.payment.paymentrecord_set.first()
            ws['J' + str(current_row)] = paid.date.strftime('%Y-%m-%d')
            if paid.method == 'cash':
                ws['K' + str(current_row)] = paid.paid
                total_paid_cash += paid.paid
            elif paid.method == 'card':
                ws['L' + str(current_row)] = paid.paid
                total_paid_transfer += paid.paid

            ws['M' + str(current_row)] = reception.payment.total - paid.paid
            total_debit += reception.payment.total - paid.paid
            ws['P' + str(current_row)] = reception.payment.progress

            ws['Q' + str(current_row)] = 'Dr.' + reception.doctor.name_short
            ws['R' + str(current_row)] = 0
            ws['S' + str(current_row)] = '=IF(R' + str(current_row) + '=0,L' + str(current_row) +',"")'
            ws['T' + str(current_row)] = reception.payment.memo

        

        data_num += 1
        form_row += 1
        current_row +=1

    begin_total_sel = current_row
    ws['A' + str(current_row)] = ''
    ws['B' + str(current_row)] = ''
    ws.merge_cells('C' + str(current_row) + ':E' + str(current_row))
    ws['C' + str(current_row)] = 'Total'


    ws['F' + str(current_row)] = total_subtotal
    ws['G' + str(current_row)] = total_discount
    ws['H' + str(current_row)] = total_total
    ws['K' + str(current_row)] = total_paid_cash
    ws['L' + str(current_row)] = total_paid_transfer
    ws['M' + str(current_row)] = total_debit

    current_row += 1
    
    rows = ws['C'+ str(current_row) + ':J' + str(current_row)]

    ws.merge_cells('C' + str(current_row) + ':E' + str(current_row))
    ws['C' + str(current_row)] = 'Total (1)'
    ws['F' + str(current_row)] = total_subtotal

    ws.merge_cells('G' + str(current_row) + ':K' + str(current_row))
    ws['G' + str(current_row)] = 'Doanh thu theo Khoa'

    
    current_row +=1
    ws.merge_cells('C' + str(current_row) + ':E' + str(current_row))
    ws['C' + str(current_row)] = 'Tổng tiền giảm giá (2)'
    ws['F' + str(current_row)] = total_discount

    ws['G' + str(current_row)] = 'Nhi'
    ws['H' + str(current_row)] = 'Da liễu'
    ws['I' + str(current_row)] = 'Nội'
    ws['J' + str(current_row)] = 'Phục hồi chức năng'
    ws['K' + str(current_row)] = 'Phẫu thuật thẩm mỹ'

    current_row +=1
    ws.merge_cells('C' + str(current_row) + ':E' + str(current_row))
    ws['C' + str(current_row)] = 'Doanh thu thực tế (1)-(2)'
    ws['F' + str(current_row)] = total_total

    derm_receptions = receptions.filter(depart = 6)
    ent_receptions= receptions.filter(depart = 5)
    ps_receptions = receptions.filter(depart = 4)
    im_receptions = receptions.filter(depart = 2)
    pm_receptions = receptions.filter(depart = 7)

    derm_total = 0
    derm_total_paid = 0
    for reception in derm_receptions:
        derm_total += reception.payment.total
        for data in reception.payment.paymentrecord_set.all():
            derm_total_paid += data.paid

    ent_total = 0
    ent_total_paid = 0
    for reception in ent_receptions:
        ent_total += reception.payment.total
        for data in reception.payment.paymentrecord_set.all():
            ent_total_paid += data.paid

    ps_total = 0
    ps_total_paid = 0
    for reception in ps_receptions:
        ps_total += reception.payment.total
        for data in reception.payment.paymentrecord_set.all():
            ps_total_paid += data.paid
    
    im_total = 0
    im_total_paid = 0
    for reception in im_receptions:
        im_total += reception.payment.total
        for data in reception.payment.paymentrecord_set.all():
            im_total_paid += data.paid

    pm_total = 0
    pm_total_paid = 0
    for reception in pm_receptions:
        pm_total += reception.payment.total
        for data in reception.payment.paymentrecord_set.all():
            pm_total_paid += data.paid

    ws['G' + str(current_row)] = derm_total
    ws['H' + str(current_row)] = ent_total
    ws['I' + str(current_row)] = ps_total
    ws['J' + str(current_row)] = im_total
    ws['K' + str(current_row)] = pm_total


    current_row +=1
    ws.merge_cells('C' + str(current_row) + ':C' + str(current_row+3))
    ws['C' + str(current_row)] = 'Phương thức \nthanh toán'
    ws.merge_cells('D' + str(current_row) + ':E' + str(current_row))
    ws['D' + str(current_row)] = 'Tiền mặt'
    ws['F' + str(current_row)] = total_paid_cash

    ws['G' + str(current_row)] = derm_total_paid
    ws['H' + str(current_row)] = ent_total_paid
    ws['I' + str(current_row)] = ps_total_paid
    ws['J' + str(current_row)] = im_total_paid
    ws['K' + str(current_row)] = pm_total_paid

    ws['L' + str(current_row)] = '=SUM(G' + str(current_row) + ':K' + str(current_row) +')-F' + str(current_row)

    current_row +=1
    ws.merge_cells('D' + str(current_row) + ':E' + str(current_row))
    ws['D' + str(current_row)] = 'VP Bank'
    ws['F' + str(current_row)] = 0

    ws['G' + str(current_row)] = 0
    ws['H' + str(current_row)] = 0
    ws['I' + str(current_row)] = 0
    ws['J' + str(current_row)] = 0
    ws['K' + str(current_row)] = 0

    current_row +=1
    ws.merge_cells('D' + str(current_row) + ':E' + str(current_row))
    ws['D' + str(current_row)] = 'BIDV'
    ws['F' + str(current_row)] = 0

    ws['G' + str(current_row)] = 0
    ws['H' + str(current_row)] = 0
    ws['I' + str(current_row)] = 0
    ws['J' + str(current_row)] = 0
    ws['K' + str(current_row)] = 0

    current_row +=1
    ws.merge_cells('D' + str(current_row) + ':E' + str(current_row))
    ws['D' + str(current_row)] = 'Chưa thanh toán'
    ws['F' + str(current_row)] = total_debit

    ws['G' + str(current_row)] = derm_total - derm_total_paid
    ws['H' + str(current_row)] = ent_total - ent_total_paid
    ws['I' + str(current_row)] = ps_total - ps_total_paid
    ws['J' + str(current_row)] = im_total - im_total_paid
    ws['K' + str(current_row)] = pm_total - pm_total_paid


    rows = ws['C'+ str(begin_total_sel) + ':K' + str(current_row)]
    for row in rows:
        for cell in row:
            cell.font = copy(style_font)
            cell.border = copy(style_border)
            cell.fill = copy(style_fill)
            cell.number_format = copy(style_number_format)
            cell.protection = copy(style_protection)
            cell.alignment = copy(style_alignment)



    #ws.cell(column= 7 , row=(begin_total_sel + 1)).style.font.bold= copy(False)
    #ws.cell(column= 3, row=(begin_total_sel + 2)).font.bold= False
    #ws.cell(column= 4, row=(begin_total_sel + 4)).font.bold= False
    #ws.cell(column= 4, row=(begin_total_sel + 5)).font.bold= False
    #ws.cell(column= 4, row=(begin_total_sel + 6)).font.bold= False
    #ws.cell(column= 4, row=(begin_total_sel + 7)).font.bold= False


    #38.25

    wb.save(response)
    return response



#고객 정보 다운로드
def cumstomer_management_excel(request):
    today = datetime.date.today().strftime('%Y%m%d')
    print(today)
    #이름 설정
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="CUSTOMER INFORMATION ' + today +'.xlsx"'

    #엑셀 파일 불러오기
    wb = load_workbook('static/excel_form/reception_report.xlsx') #Workbook()
    ws = wb.active# grab the active worksheet


    wb.save(response)
    return response


def search_medicine(request):
    filter = request.POST.get('filter')
    string = request.POST.get('string')

    date_start = request.POST.get('start_end_date').split(' - ')[0]
    date_end = request.POST.get('start_end_date').split(' - ')[1]

    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end, "%Y-%m-%d").date(), datetime.time.max)

    
    #if filter != '':
    #    kwargs.update({'depart_id':depart})


    datas_dict= {}
    receptions = Reception.objects.filter(recorded_date__range = (date_min, date_max), progress = 'done').order_by("-id")
    for reception in receptions:
        if hasattr(reception,'diagnosis'):
            medicine_manager_set = MedicineManager.objects.filter(diagnosis_id = reception.diagnosis.id)
            for medicine_manager in medicine_manager_set:
                medicine = Medicine.objects.get(pk = medicine_manager.medicine_id)
            
                if medicine.code in datas_dict.keys():
                    datas_dict[medicine.code]['sales'] += medicine_manager.amount * medicine_manager.days
                    datas_dict[medicine.code]['total_salse'] += medicine_manager.amount * medicine_manager.days * medicine.get_price(reception.recorded_date)

                else:
                    datas_dict.update({
                        medicine.code:{
                            'code':medicine.code,
                            'name':medicine.name,
                            'ingredient':'' if medicine.ingredient is None else medicine.ingredient,
                            'company':'' if medicine.company is None else medicine.company,
                            'count':medicine.inventory_count,
                            'price':medicine.get_price(reception.recorded_date),
                            'sales' : medicine_manager.amount * medicine_manager.days,
                            'total_salse' : medicine_manager.amount * medicine_manager.days * medicine.get_price(reception.recorded_date)
                            }
                        })

        
    id = 1
    total_amount = 0
    datas=[]
    datas_dict_sorted = sorted(datas_dict.items())
    for data_dict in datas_dict_sorted:
        data_dict[1].update({
            'id':id,
            })
        datas.append(data_dict)
        total_amount += data_dict[1]['total_salse']
        id += 1

    page = request.POST.get('page',1)
    page_context = request.POST.get('context_in_page',10)

    paginator = Paginator(datas, page_context)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)
    

    context = {
            'datas':list(paging_data),
            'total_amount':total_amount,

            'page_range_start':paging_data.paginator.page_range.start,
            'page_range_stop':paging_data.paginator.page_range.stop,
            'page_number':paging_data.number,
            'has_previous':paging_data.has_previous(),
            'has_next':paging_data.has_next(),
        }

    return JsonResponse(context)


@login_required
def temp_doctor_audit(request):
    doctor_search_form = DoctorsSearchForm()

    
    list_exam_fee = []
    list_precedures = []
    list_radiologys = []


    exam_fees = ExamFee.objects.filter(Q(code = 'E0010') | Q(code = 'E0011'))
    for exam_fee in exam_fees:
        list_exam_fee.append({'code':exam_fee.code,'value':exam_fee.name})

    precedures = Precedure.objects.filter(code__contains='PM')
    for precedure in precedures:
        list_precedures.append({'code':precedure.code,'value':precedure.name})

    radiologys = Precedure.objects.filter(code__contains='R', precedure_class_id = 10 )
    for radiology in radiologys:
        list_radiologys.append({'code':radiology.code,'value':radiology.name})

    return render(request,
    'Doctor/audit_PM.html',
        {
            'doctor_search':doctor_search_form,

            'list_exam_fee':list_exam_fee,
            'list_precedures':list_precedures,
            'list_radiologys':list_radiologys,

        }
    )


def search_patient(request):


    return JsonResponse(context)



def inventory_test(request):
    class_datas=[]
    test_class = TestClass.objects.all()
    for tmp_class in test_class:
        class_datas.append({
            'id':tmp_class.id,
            'name': tmp_class.get_name_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
            })


    return render(request,
    'Manage/inventory_test.html',
            {
                'test_class':class_datas,
            },
        )



def test_search(request):
    string = request.POST.get('string')
    filter = request.POST.get('filter')
    class_id = request.POST.get('class_id')

    kwargs = {}
    if class_id != '':
        kwargs.update({
            'test_class_id':class_id,
            })

    argument_list = [] 
    if string !='':
        argument_list.append( Q(**{'name__icontains':string} ) )
        argument_list.append( Q(**{'name_vie__icontains':string} ) )
   


    if string == '' :
        query_datas = Test.objects.filter(**kwargs).select_related('test_class').exclude(use_yn = 'N').order_by("name")
    #elif filter == 'name':
    #    query_datas = Test.objects.filter( Q(name__icontains = string) | Q(name_vie__icontains = string)).select_related('test_class').exclude(use_yn = 'N').order_by("name")
    else:
        query_datas = Test.objects.filter(functools.reduce(operator.or_, argument_list),**kwargs).select_related('test_class').exclude(use_yn = 'N').order_by("name")


    datas=[]
    for query_data in query_datas:
        data = {
                'id' : query_data.id,
                'code': query_data.code,
                'name' : query_data.get_name_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
                'class':query_data.test_class.name,
                'price' : query_data.get_price(),
                'price_dollar' : query_data.get_price_dollar(),
            }
        datas.append(data)


    page = request.POST.get('page',1)
    context_in_page = request.POST.get('context_in_page');
    paginator = Paginator(datas, context_in_page)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)

    context = {
        #'datas':datas,
        'datas':list(paging_data),
        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),
        
        }
    return JsonResponse(context)


def test_add_edit_get(request):
    id = request.POST.get('id');

    test = Test.objects.get(id=id)


    return JsonResponse({
        'id':test.id,
        'name':test.name,
        'name_vie':test.name_vie,
        'price':test.get_price(),
        'price_dollar':test.get_price_dollar(),
        'precedure_class_id':test.test_class_id,
        'result':True,
        })


def test_add_edit_set(request):
    id = request.POST.get('id');
    type = request.POST.get('type');
    test_class = request.POST.get('test_class');
    name = request.POST.get('name');
    name_vie = request.POST.get('name_vie');
    price = request.POST.get('price');
    price_dollar = request.POST.get('price_dollar');

    if int(id) == 0 :
        data = Test()
        data.price = price
        

        last_code = Test.objects.last()
        test_class = int(test_class)
        CODE = 'L'
            
        if last_code == None:
           data.code = CODE + str('0001')
        else:
            temp_code = last_code.code.split(CODE)
            data.code = CODE + str('%04d' % (int(temp_code[1]) + 1))
    else:
        data = Test.objects.get(id=id)
        now = datetime.datetime.now()
        now = now - datetime.timedelta(seconds = 1) 
        str_now = now.strftime('%Y%m%d%H%M%S')
        try: 
            old_price = Pricechange.objects.get(type="Test",country='VI',type2='OUTPUT',code=data.code, date_end="99999999999999")
            
            if old_price.price != int(price):
                old_price.date_end = str_now
                old_price.save()

                new_price = Pricechange(type="Test",country='VI',type2='OUTPUT',code=data.code)
                new_price.price = price
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if data.price != int(price):
                new_price = Pricechange(type="Test",country='VI',type2='OUTPUT',code=data.code)
                new_price.price = price
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()
                
        try:
            old_price_dollar = Pricechange.objects.get(type="Test",country='US',type2='OUTPUT',code=data.code, date_end="99999999999999")
            
            if old_price_dollar.price != int(price_dollar):
                old_price_dollar.date_end = str_now
                old_price_dollar.save()

                new_price = Pricechange(type="Test",country='US',type2='OUTPUT',code=data.code)
                new_price.price = price_dollar
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if data.price != int(price_dollar):
                new_price = Pricechange(type="Test",country='US',type2='OUTPUT',code=data.code)
                new_price.price = price_dollar
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

    data.test_class_id = test_class
    data.name = name
    data.name_vie = name_vie
    data.save()


    return JsonResponse({
        'result':True,
        })



    return JsonResponse({})

def test_add_edit_delete(request):

    id = request.POST.get('id')
    test = Test.objects.get(id=id)
    test.use_yn = 'N'
    test.save()

    return JsonResponse({
        'result':True,
        })


def inventory_precedure(request):

    class_datas=[]
    precedure_class = PrecedureClass.objects.all()
    for tmp_class in precedure_class:
        class_datas.append({
            'id':tmp_class.id,
            'name': tmp_class.get_name_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
            })


    return render(request,
    'Manage/inventory_precedure.html',
            {
                'precedure_class':class_datas,
            },
        )


def precedure_search(request):
    string = request.POST.get('string')
    filter = request.POST.get('filter')
    class_id = request.POST.get('class_id')

    kwargs = {}
    if class_id != '':
        kwargs.update({
            'precedure_class_id':class_id,
            })

    argument_list = [] 
    if string !='':
        argument_list.append( Q(**{'name__icontains':string} ) )
        argument_list.append( Q(**{'name_vie__icontains':string} ) )
   


    if string == '' :
        query_datas = Precedure.objects.filter(**kwargs).select_related('precedure_class').exclude(use_yn = 'N').order_by("name")
    #elif filter == 'name':
    #    query_datas = Test.objects.filter( Q(name__icontains = string) | Q(name_vie__icontains = string)).select_related('test_class').exclude(use_yn = 'N').order_by("name")
    else:
        query_datas = Precedure.objects.filter(functools.reduce(operator.or_, argument_list),**kwargs).select_related('precedure_class').exclude(use_yn = 'N').order_by("name")



    datas=[]
    for query_data in query_datas:
        data = {
                'id' : query_data.id,
                'code': query_data.code,
                'name' : query_data.get_name_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
                'class':query_data.precedure_class.name,
                'price' : query_data.get_price(),
            }

        datas.append(data)


    page = request.POST.get('page',1)
    context_in_page = request.POST.get('context_in_page');
    paginator = Paginator(datas, context_in_page)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)

    context = {
        #'datas':datas,
        'datas':list(paging_data),
        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),
        
        }
    return JsonResponse(context)

def precedure_add_edit_get(request):
    id = request.POST.get('id');

    precedure = Precedure.objects.get(id=id)


    return JsonResponse({
        'id':precedure.id,
        'name':precedure.name,
        'name_vie':precedure.name_vie,
        'price':precedure.get_price(),
        'price_dollar':precedure.get_price_dollar(),
        'precedure_class_id':precedure.precedure_class_id,
        'result':True,
        })


def precedure_add_edit_set(request):
    id = request.POST.get('id');
    type = request.POST.get('type');
    precedure_class = request.POST.get('precedure_class');
    name = request.POST.get('name');
    name_vie = request.POST.get('name_vie');
    price = request.POST.get('price');
    price_dollar = request.POST.get('price_dollar');

    if int(id) == 0 :
        data = Precedure()
        data.price = price
        data.price_dollar = price_dollar

        last_code = Precedure.objects.filter(precedure_class_id=precedure_class).order_by('code').last()
        precedure_class = int(precedure_class)

        if precedure_class ==1: #D
            CODE = 'D'
        elif precedure_class == 2: #CT
            CODE = 'CT'
        elif precedure_class == 3: #ENT
            CODE = 'E'
        elif precedure_class == 4: #GE
            CODE = 'GE'
        elif precedure_class == 5: #Radi
            CODE = 'R'
        elif precedure_class == 6: #U
            CODE = 'U'
        elif precedure_class == 7: #P
            CODE = 'P'
        elif precedure_class == 8: #T
            CODE = 'T'
        elif precedure_class == 10: #PM
            CODE = 'PM'
        elif ( precedure_class >= 11 and precedure_class <= 30 ) or precedure_class == 9: #: #DERM
            CODE = 'DM'
        elif ( precedure_class >= 31 and precedure_class <=40 ) or precedure_class == 42: #PS
            CODE = 'PS'
        elif precedure_class == 41: #MRI
            CODE = 'MRI'
        elif precedure_class == 44: #IM
            CODE = 'IM'
        elif precedure_class == 45: #Vaccin
            CODE = 'VC'
        elif precedure_class == 46: #Vaccin
            CODE = 'EC'

        if last_code == None:
           data.code = CODE + str('0001')
        else:
            temp_code = last_code.code.split(CODE)
            data.code = CODE + str('%04d' % (int(temp_code[1]) + 1))
        

    else:
        data = Precedure.objects.get(id=id)
        str_now = datetime.datetime.now().strftime('%Y%m%d%H%M%S')

        try: 
            old_price = Pricechange.objects.get(type="Precedure",country='VI',type2='OUTPUT',code=data.code, date_end="99999999999999")
            
            if old_price.price != int(price):
                old_price.date_end = str_now
                old_price.save()

                new_price = Pricechange(type="Precedure",country='VI',type2='OUTPUT',code=data.code)
                new_price.price = price
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if data.price != int(price):
                new_price = Pricechange(type="Precedure",country='VI',type2='OUTPUT',code=data.code)
                new_price.price = price
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()
                
        try:
            old_price_dollar = Pricechange.objects.get(type="Precedure",country='US',type2='OUTPUT',code=data.code, date_end="99999999999999")
            
            if old_price_dollar.price != int(price_dollar):
                old_price_dollar.date_end = str_now
                old_price_dollar.save()

                new_price = Pricechange(type="Precedure",country='US',type2='OUTPUT',code=data.code)
                new_price.price = price_dollar
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if data.price != int(price_dollar):
                new_price = Pricechange(type="Precedure",country='US',type2='OUTPUT',code=data.code)
                new_price.price = price_dollar
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

    data.precedure_class_id = precedure_class
    data.name = name
    data.name_vie = name_vie
    data.save()


    return JsonResponse({
        'result':True,
        })


def precedure_add_edit_delete(request):
    id = request.POST.get('id')
    precedure = Precedure.objects.get(id=id)
    precedure.use_yn = 'N'
    precedure.save()

    return JsonResponse({
        'result':True,
        })




def medicine_search(request):
    string = request.POST.get('string')
    filter = request.POST.get('filter')
    class_id = request.POST.get('class_id')
    

    kwargs = {}
    if class_id != '':
        kwargs.update({
            'medicine_class_id':class_id,
            'type':"TOOL",
            })

    
    datas=[]
    if string == '':
        argument_list_expriry_date = []
        argument_list_expriry_date.append( Q(**{'medicine__code__icontains':'T'} ) )

        expiry_date = datetime.datetime.now() + datetime.timedelta(days=180)
        #medicine_tmp = MedicineLog.objects.filter(type='add',expiry_date__lte = expiry_date ).select_related('medicine').exclude(medicine__use_yn='N' ,expiry_date=None,tmp_count=0).order_by('expiry_date')
        medicine_tmp= MedicineLog.objects.filter( medicine__type="TOOL", type='add', expiry_date__lte = expiry_date, tmp_count__gte= 0 ).select_related('medicine').values(
            'medicine_id',
            ).annotate(Count('medicine_id')).order_by('expiry_date')
       
        #medicine_tmp = MedicineLog.objects.filter( type='add', expiry_date__lte = expiry_date).exclude(tmp_count__lt = 0,medicine__use_yn='N',expiry_date=None).order_by('expiry_date').select_related('medicine')
        for tmp in medicine_tmp:
            medicine = Medicine.objects.get(id = tmp['medicine_id'])
            data = {
                    'id' : medicine.id,
                    'code': medicine.code,
                    'name' : medicine.name,
                    'company' : '' if medicine.company is None else medicine.company,
                    'unit' : '' if medicine.get_unit_lang(request.session[translation.LANGUAGE_SESSION_KEY]) is None else medicine.get_unit_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
                    'price' : medicine.get_price(),
                    'count' : medicine.inventory_count,
                    'alaert_expiry':True,
                }
            datas.append(data)
            
        medicines = Medicine.objects.filter(**kwargs, type="TOOL").exclude(use_yn='N').order_by('name')
    else:
        argument_list = [] 
        #argument_list = [] 
        argument_list.append( Q(**{'name__icontains':string} ) )
        argument_list.append( Q(**{'name_vie__icontains':string} ) )
        argument_list.append( Q(**{'name_display__icontains':string} ) )

        medicines = Medicine.objects.filter(functools.reduce(operator.or_, argument_list),**kwargs,type="TOOL").exclude(use_yn='N').order_by("name")#.select_related('medicine_class').exclude(use_yn = 'N').order_by("name")



    
    for medicine in medicines:
        data = {
                'id' : medicine.id,
                'code': medicine.code,
                'name' : medicine.name,
                'company' : '' if medicine.company is None else medicine.company,
                'unit' : '' if medicine.get_unit_lang(request.session[translation.LANGUAGE_SESSION_KEY]) is None else medicine.get_unit_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
                'price' : medicine.get_price(),
                'count' : medicine.inventory_count,
            }

        datas.append(data)


    page = request.POST.get('page',1)
    context_in_page = request.POST.get('context_in_page');
    paginator = Paginator(datas, context_in_page)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)


    

    context = {
        #'datas':datas,
        'datas':list(paging_data),
        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),
        
        }
    return JsonResponse(context)


def set_data_control(request):
    medicine_id = request.POST.get('medicine_id');


    medicine = Medicine.objects.get(pk = medicine_id)

    context = {
        'name':medicine.name,
        'company':medicine.company,
        'country':medicine.country,
        'ingredient':medicine.ingredient,
        'unit':medicine.unit,
        'price':medicine.get_price(),
        }
    return JsonResponse(context)


def save_data_control(request):
    selected_option = request.POST.get('selected_option');
    name = request.POST.get('name');
    price = request.POST.get('price');
    company = request.POST.get('company');
    ingredient = request.POST.get('ingredient');
    unit = request.POST.get('unit');
    changes = request.POST.get('changes') if request.POST.get('changes') is not '' else 0;


    if selected_option == 'new':
        medicine = Medicine()
        log = MedicineLog(type='new')
        code = Medicine.objects.all().last()
        code = 'M{:04d}'.format(code.id)
        medicine.price = price
    else:
        medicine = Medicine.objects.get(pk = selected_option )
        if int(changes) < 0:
            log = MedicineLog(type='dec')
        else:
            log = MedicineLog(type='add')

    medicine.name = name
    #medicine.price = price
    medicine.company = company
    medicine.ingredient = ingredient
    medicine.unit = unit

    medicine.inventory_count += int(changes)
    medicine.save()
    log.medicine = medicine
    log.changes = int(changes)
    log.save()
     
    context = {'result':True}
    return JsonResponse(context)

def inventory_medical_tool(request):
    medicine_search_form = MedicineSearchForm()

    price_multiple_level = COMMCODE.objects.filter(commcode_grp = 'MED_MULTI_CODE').values('commcode','se1','se2').order_by('commcode_grp')


    if request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        medicine_class = MedicineClass.objects.all().annotate(name_display = F('name_vie')).values('id','name_display')
    else:
        medicine_class = MedicineClass.objects.all().annotate(name_display = F('name')).values('id','name_display')

    type = ["Medical Tool"]
    

    return render(request,
    'Manage/inventory_medical_tool.html',
            {
                'medicinesearch':medicine_search_form,
                'price_multiple_level':price_multiple_level,
                'medicine_class':medicine_class,
                'type':type,

            },
        )


def medicine_add_edit_get(request):
    id = request.POST.get('id')

    try:
        medicine = Medicine.objects.get(id=id)

        context = {
            'result':True,
            'id':medicine.id,
            'code':medicine.code,
            'name':medicine.name,
            'name_vie':medicine.name_vie,
            'unit':medicine.unit,
            'unit_vie':medicine.unit_vie,
            'company':medicine.company,

            'type':'Medical Tool' if 'T' in medicine.code else 'Medical Tool',

            'price':medicine.get_price(),
            'price_input':medicine.get_price_input(),
            'price_dollar':medicine.get_price_dollar(),
            'multiple_level':medicine.multiple_level,

            'inventory_count':medicine.inventory_count,
            'medicine_class_id':medicine.medicine_class_id,

            }
    except Medicine.DoesNotExist:
        context = {'result':False}




    return JsonResponse(context)


def medicine_add_edit_set(request):
    id = int(request.POST.get('id'))
    code = request.POST.get('code')
    type = request.POST.get('type')
    medicine_class = request.POST.get('medicine_class')
    name = request.POST.get('name')
    name_vie = request.POST.get('name_vie')
    unit = request.POST.get('unit')
    unit_vie = request.POST.get('unit_vie')
    company = request.POST.get('company')
    price_input = request.POST.get('price_input')
    multiple_level = request.POST.get('multiple_level')
    price = request.POST.get('price')   
    price_dollar = request.POST.get('price_dollar')

    if int(id) == 0 :
        data = Medicine()
        if type in 'Medical Tool':
            last_code =Medicine.objects.filter(code__icontains="T").last()
            temp_code = last_code.code.split('T')
            code = 'T' + str('%04d' % (int(temp_code[1]) + 1))

        data.code = code
        data.price = int(price)
        data.price_input = int(price_input)
        data.price_dollar = int(price_dollar)
        data.type = 'TOOL'
       
    else:
        data = Medicine.objects.get(id=id)
        str_now = datetime.datetime.now().strftime('%Y%m%d%H%M%S')

        try: 
            old_price = Pricechange.objects.get(type="Medicine",country='VI',type2='OUTPUT',code=data.code, date_end="99999999999999")
            
            if old_price.price != int(price):
                old_price.date_end = str_now
                old_price.save()

                new_price = Pricechange(type="Medicine",country='VI',type2='OUTPUT',code=data.code)
                new_price.price = price
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if data.price != int(price):
                new_price = Pricechange(type="Medicine",country='VI',type2='OUTPUT',code=data.code)
                new_price.price = price
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()


        try:
            old_price_input = Pricechange.objects.get(type="Medicine",country='VI',type2='INPUT',code=data.code, date_end="99999999999999")
            
            if old_price_input.price != int(price_input):
                old_price_input.date_end = str_now
                old_price_input.save()

                new_price = Pricechange(type="Medicine",country='VI',type2='INPUT',code=data.code)
                new_price.price = price_input
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if data.price_input != int(price_input):
                new_price = Pricechange(type="Medicine",country='VI',type2='INPUT',code=data.code)
                new_price.price = price_input
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        try:
            old_price_dollar = Pricechange.objects.get(type="Medicine",country='US',type2='OUTPUT',code=data.code, date_end="99999999999999")
            
            if old_price_dollar.price != int(price_dollar):
                old_price_dollar.date_end = str_now
                old_price_dollar.save()

                new_price = Pricechange(type="Medicine",country='US',type2='OUTPUT',code=data.code)
                new_price.price = price_dollar
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if data.price_dollar != int(price_dollar):
                new_price = Pricechange(type="Medicine",country='US',type2='OUTPUT',code=data.code)
                new_price.price = price_dollar
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()


    data.medicine_class_id = medicine_class
    data.name = name
    data.name_vie = name_vie
    data.unit = unit
    data.unit_vie = unit_vie
    data.company = company
    data.multiple_level = multiple_level
    

    data.save()

    context = {'result':True}




    return JsonResponse(context)


def medicine_add_edit_check_code(request):
    code = request.POST.get('code')
    id = request.POST.get('id')
    try:
        medicine = Medicine.objects.get(code=code)
        res = "N"
        
        if medicine.id == int(id):
            res = 'Same'
    except Medicine.DoesNotExist:
        res = "Y"

    context = {'result':res}

    return JsonResponse(context)


def medicine_add_edit_delete(request):
    id = request.POST.get('id')

    medicine = Medicine.objects.get(id=id)
    medicine.use_yn = 'N'
    medicine.save()

    log = MedicineLog()
    log.medicine = medicine
    log.type='del'
    log.save()



    return JsonResponse({'result':True})

def get_inventory_history(request):
    id = request.POST.get('id')
    date = request.POST.get('date')


    #date_min = datetime.datetime.combine(datetime.datetime.strptime(date, "%Y-%m-%d").date(), datetime.time.min)
    #date_max = datetime.datetime.combine(datetime.datetime.strptime(date, "%Y-%m-%d").date(), datetime.time.max)

    #medicine_logs = MedicineLog.objects.filter(date__range = (date_min, date_max), medicine_id = id).values('date','changes','type','memo').order_by('-date')
    medicine_logs = MedicineLog.objects.filter(medicine_id = id).values('date','changes','type','memo').order_by('-date')
    datas = []
    for medicine_log in medicine_logs:
        data = {
            'date':medicine_log['date'].strftime('%Y-%m-%d %H:%M:%S'),
            'changes':medicine_log['changes'],
            'type':medicine_log['type'],
            'memo':'' if medicine_log['memo'] is None else medicine_log['memo'],
            }
        datas.append(data)

    medicine = Medicine.objects.get(id=id)
    return JsonResponse({
        'datas':datas,
        'count':medicine.inventory_count,
        })

def save_database_add_medicine(request):
    id = request.POST.get('id')
    registration_date= request.POST.get('registration_date')
    expiry_date= request.POST.get('expiry_date')
    changes= request.POST.get('changes')
    memo= request.POST.get('memo')
    check= request.POST.get('check')
    
    if int(check)==0 :
        medicine = Medicine.objects.get(id=id)
        count = medicine.inventory_count
        medicine.inventory_count =  count + int(changes)
        medicine.save()

        medicine_logs = MedicineLog()
        medicine_logs.changes = changes
        medicine_logs.medicine = medicine
    else:
        medicine_logs = MedicineLog.objects.get(id = check)
        medicine_logs.date = datetime.datetime.strptime(registration_date, '%Y-%m-%d')

    medicine_logs.memo = memo
    if expiry_date != '' :
        medicine_logs.expiry_date = datetime.datetime.strptime(expiry_date, '%Y-%m-%d')
    
 
    medicine_logs.tmp_count = changes
    medicine_logs.type='add'
    medicine_logs.save()




    input_price = request.POST.get('input_price',None)
    if input_price:
        try:
            str_now = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
            old_price_input = Pricechange.objects.get(type="Medicine",country='VI',type2='INPUT',code=medicine.code, date_end="99999999999999")
            
            if old_price_input.price != int(input_price):
                old_price_input.date_end = str_now
                old_price_input.save()

                new_price = Pricechange(type="Medicine",country='VI',type2='INPUT',code=medicine.code)
                new_price.price = price_input
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if medicine.price_input != int(input_price):
                new_price = Pricechange(type="Medicine",country='VI',type2='INPUT',code=medicine.code)
                new_price.price = input_price
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

    return JsonResponse({'result':True})


def get_expiry_date(request):

    id = request.POST.get('id')
    medicine = Medicine.objects.get(id=id)

    medicine_logs = MedicineLog.objects.filter(medicine = medicine, type='add').exclude(tmp_count__lte = 0).order_by('expiry_date').values('id','date','tmp_count','expiry_date')
    
    datas = []
    for medicine_log in medicine_logs:
        
        data = {
            'id':medicine_log['id'],
            'date':medicine_log['date'].strftime('%Y-%m-%d'),
            'expiry_date':0 if medicine_log['expiry_date'] is None else medicine_log['expiry_date'].strftime('%Y-%m-%d'),
            'tmp_count':medicine_log['tmp_count'] if medicine_log['tmp_count'] is not None else 0,
            }
        datas.append(data)

    return JsonResponse({
        'result':True,
        'datas':datas,
        })


def get_edit_database_add_medicine(request):
    id = request.POST.get('id')

    data = MedicineLog.objects.values('id','date','expiry_date','memo','tmp_count').get(id=id)

    return JsonResponse({
        'result':True,
        'data':{
            'id':data['id'],
            'date':data['date'].strftime('%Y-%m-%d'),
            'expiry_date':'' if data['expiry_date'] is None else data['expiry_date'].strftime('%Y-%m-%d'),
            'memo':data['memo'],
            'tmp_count':data['tmp_count'],
            },
        })


def save_database_disposal_medicine(request):
    id = request.POST.get('id')
    disposial = request.POST.get('disposial')
    memo = request.POST.get('memo')


    disposal_data = MedicineLog.objects.get(id=id)
    if disposal_data.tmp_count < int(disposial):
        return JsonResponse({
            'result':False,
            'msg':1, # tmp 카운트가 더 적을 경우
                             })
    else:
        medicine = Medicine.objects.get(id = disposal_data.medicine_id)
        disposal_data.tmp_count -= int(disposial)
        
        disposal_data.save()

        medicine_log = MedicineLog()
        medicine_log.type = 'dec'
        medicine_log.changes = disposial
        medicine_log.memo = memo
        medicine_log.medicine_id = disposal_data.medicine_id
        medicine_log.save()

        
        medicine.inventory_count -= int(disposial)
        medicine.save()



    return JsonResponse({'result':True,})



#기안서
def draft(request):

    if request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    else:
        f_name = F('commcode_name_en')


    if request.META['SERVER_PORT'] == '9090' or request.META['SERVER_PORT'] == '11111':#테스트서버
        url = 'Manage/Draft.html'
        #부서 - 병원
        depart_type = COMMCODE.objects.filter(upper_commcode = '000002',commcode_grp = 'DEPART_CLICINC',use_yn="Y").annotate(code = F('commcode'),name = f_name ).values('code','name','id')

    elif request.META['SERVER_PORT'] == '8888':#경천애인
        url = 'Manage/Draft_KBL.html'
        #부서 - KBL
        depart_type= COMMCODE.objects.filter(upper_commcode = '000002',commcode_grp = 'DEPART_KBL', use_yn="Y").annotate(code = F('commcode'),name = f_name ).values('code','name','id')



    #기안서 종류
    draft_type = COMMCODE.objects.filter(upper_commcode = '000007',commcode_grp = 'DRAFT_TYPE',use_yn="Y").annotate(code = F('commcode'),name = f_name ).values('code','name').order_by('seq')

    #기안서 상태
    draft_status = COMMCODE.objects.filter(upper_commcode = '000007',commcode_grp = 'DRAFT_STATUS',use_yn="Y").annotate(code = F('commcode'),name = f_name ).values('code','name','id')

    file_form = board_file_form()

    return render(request,
        url,
            {
                'draft_type':draft_type,

                'depart_type':depart_type,

                'draft_status':draft_status,

                'file_form':file_form,
            }
        )


#기안서 검색
def draft_search(request):
    start = request.POST.get('start','')
    end = request.POST.get('end','')


    string = request.POST.get('string','')

    type = request.POST.get('type','')
    requester = request.POST.get('requester','')
    status = request.POST.get('status','')

    if request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
        f_user = F('name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
        f_user = F('name_ko')
    else:
        f_name = F('commcode_name_en')
        f_user = F('name_en')

    date_min = datetime.datetime.combine(datetime.datetime.strptime(start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(end, "%Y-%m-%d").date(), datetime.time.max)


    kwargs={}
    if string != '':
        kwargs['string__icontains']=string
    if type != '':
        kwargs['type']=type
    if requester != '':
        kwargs['requester']=requester
    if status != '':
        kwargs['status']=status

    print(kwargs)

    draft_list=[]
    if request.META['SERVER_PORT'] == '8888':#경천애인
        draft_query = Draft.objects.filter(date_registered__range = (date_min, date_max),**kwargs,use_yn='Y',is_KBL='Y').order_by('-date_registered')
    else:
        draft_query = Draft.objects.filter(date_registered__range = (date_min, date_max),**kwargs,use_yn='Y',is_KBL='N').order_by('-date_registered')

    for draft in draft_query:
        depart = COMMCODE.objects.filter(id = draft.depart).annotate(name = f_name ).values('name')[:1]
        status_val = COMMCODE.objects.filter(upper_commcode = '000007',commcode_grp='DRAFT_STATUS', commcode=draft.status).annotate(name = f_name ).values('name')[:1]

        user_creator = User.objects.filter(id = draft.creator).annotate(name = f_user ).values('name')[:1]

        
        if draft.date_in_charge == '0000-00-00 00:00:00':
            in_charge = None
        else:
            in_charge = draft.date_in_charge[0:10]

        if draft.date_leader == '0000-00-00 00:00:00':
            leader = None
        else:
            leader = draft.date_leader[0:10]

        if draft.date_accounting == '0000-00-00 00:00:00':
            accounting = None
        else:
            accounting = draft.date_accounting[0:10]

        if draft.date_ceo == '0000-00-00 00:00:00':
            ceo = None
        else:
            ceo = draft.date_ceo[0:10]


        

        draft_list.append({
            'id':draft.id,
            'status':draft.status,
            'type':draft.type,
            'title':draft.title,
            'depart':depart[0]['name'],
            'RQSTR':user_creator[0]['name'],
            'RQSTD_DATE':draft.date_registered[0:10],
            'in_charge':in_charge,
            'leader':leader,
            'accounting':accounting,
            'ceo': ceo,
            
            'status_val':status_val[0]['name'],
            })

    page = request.POST.get('page',1)
    context_in_page = request.POST.get('page_context');
    paginator = Paginator(draft_list, context_in_page)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)


    return JsonResponse({

        'datas':list(paging_data),
        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),
        })


#기안서 불러오기
def draft_get_data(request):
    form_id = request.POST.get('id')

    query_data = Draft.objects.get(id = form_id)



    return JsonResponse({
        'result':True,

        'type':query_data.type,
        'depart':query_data.depart,
        'creator':query_data.creator,
        'title':query_data.title,
        'contents':query_data.contents,
        'consultation':query_data.consultation,
        'additional':query_data.additional,
        'status':query_data.status,
        })






#기안서 양식
def draft_get_form(request):

    form_id = request.POST.get('form_id')

    path = 'static/draft/' + form_id
    file = open(path,'rt',encoding='UTF-8')
    data = file.read()

    return JsonResponse({'result':True,
                         'data':data,
                         })

#기안서 저장
def draft_save(request):

    id = request.POST.get("id",'')

    new_edit_type=request.POST.get("new_edit_type")
    new_edit_depart=request.POST.get("new_edit_depart")
    new_edit_name=request.POST.get("new_edit_name")
    new_edit_title=request.POST.get("new_edit_title")
    new_edit_content=request.POST.get("new_edit_content")
    new_edit_consultation=request.POST.get("new_edit_consultation")
    new_edit_MORE_CMNTS=request.POST.get("new_edit_MORE_CMNTS")
    new_edit_status=request.POST.get("new_edit_status")

    print(id)

    if id != '':
        draft = Draft.objects.get(pk = id)
    else:
        draft = Draft()
        draft.creator = request.user.id
        draft.date_registered = datetime.datetime.now()

    

    draft.type = new_edit_type
    draft.depart = new_edit_depart
    draft.title = new_edit_title
    draft.request_user = new_edit_name
    draft.contents = new_edit_content
    draft.consultation = new_edit_consultation
    draft.additional = new_edit_MORE_CMNTS
    draft.status = new_edit_status

    
    draft.modifier = request.user.id
    draft.date_last_modified = datetime.datetime.now()

    if request.META['SERVER_PORT'] == '8888':#경천애인
        draft.is_KBL='Y'

    draft.save()

    return JsonResponse({'result':True,})


#기안서 삭제
def draft_delete(request):
    id = request.POST.get('id')

    draft = Draft.objects.get(id = id )
    draft.use_yn = "N"
    draft.save()



    return JsonResponse({'result':True,})


#기안서 파일 리스트
def draft_list_file(request):
    id = request.POST.get('id')

    list_file = []
    query_file = Board_File.objects.filter(board_id = id,board_type='DRAFT').order_by('registered_date')

    for file in query_file:
        list_file.append({
            'id':file.id,
            'url':file.file.url,
            'name':file.title,
            'origin_name':file.origin_name,
            'date':file.registered_date.strftime("%Y-%m-%d"),
            'creator':file.user,
            'memo':file.memo,
            });

    return JsonResponse({
        'result':True,
        'datas':list_file,
        })

#기안서 파일 정보 불러오기
def draft_get_file(request):
    id = request.POST.get('id')

    query_file = Board_File.objects.get(id = id)

    file_name = query_file.file.url,

    return JsonResponse({
        'result':True,
        'title':query_file.title,
        'memo':query_file.memo,
        'origin_name':query_file.origin_name,
        })

#기안서 파일 저장
def draft_save_file(request):
    id = request.POST.get('id')

    if request.method == 'POST':

        selected_file_id = request.POST.get('selected_file_id','')#파일 ID
        selected_file_list = request.POST.get('selected_file_list','')#기안서 ID 
        new_edit_file_name = request.POST.get('new_edit_file_name','')#문서 이름
        new_edit_file_remark = request.POST.get('new_edit_file_remark','')#문서 설명

        form = board_file_form(request.POST, request.FILES)
        files = request.FILES.getlist('file') 

        if form.is_valid():
           print(selected_file_id)
           if selected_file_id != '': #수정
                file_instance = Board_File.objects.get(id = selected_file_id)
           else:
                file_instance = Board_File()


           #    old_data
           #file save
           for f in files:
               if file_instance.file:
                    if os.path.isfile(file_instance.file.path):
                        os.remove(file_instance.file.path)

               file_instance.file = f
               file_instance.origin_name = f._name
               pass

           file_instance.board_id = selected_file_list
           file_instance.user = request.user.user_id
           file_instance.board_type = 'DRAFT'
           file_instance.title = new_edit_file_name
           file_instance.memo = new_edit_file_remark
           file_instance.save()
           #form.save()
           return JsonResponse({'error': False, 'message': 'Uploaded Successfully'})
        else:
           return JsonResponse({'error': True, 'errors': form.errors})

    else:
        form = board_file_form()
        return render(request, 'django_image_upload_ajax.html', {'form': form})



#기안서 파일 삭제
def draft_delete_file(request):

    id = request.POST.get('id')
    Board_File.objects.get(id=id).delete()



    return JsonResponse({
        'result':True,
        })



#기안서 승인
def check_appraove(request):

    id = request.POST.get('id')
    type = request.POST.get('type')
    val = request.POST.get('val')

    print(val)

    draft = Draft.objects.get(id = id )

    if val=='true':
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
        user_id = request.user.id
        name_en = request.user.name_en
        name_ko = request.user.name_ko
        name_vi = request.user.name_vi
    else:
        now = '0000-00-00 00:00:00'
        user_id = ''
        name_en = ''
        name_ko = ''
        name_vi = ''
    


    if type == 'incharge':
        draft.date_in_charge = now
        draft.user_id_in_charge = user_id
        draft.name_en_in_charge = name_en
        draft.name_ko_in_charge = name_ko
        draft.name_vi_in_charge = name_vi


    elif type == 'leader':
        draft.date_leader = now
        draft.user_id_leader= user_id
        draft.name_en_leader= name_en
        draft.name_ko_leader= name_ko
        draft.name_vi_leader= name_vi



    elif type == 'accounting':
        draft.date_accounting = now
        draft.user_id_accounting= user_id
        draft.name_en_accounting= name_en
        draft.name_ko_accounting= name_ko
        draft.name_vi_accounting= name_vi

    elif type == 'ceo':
        draft.date_ceo = now
        draft.user_id_ceo= user_id
        draft.name_en_ceo= name_en
        draft.name_ko_ceo= name_ko
        draft.name_vi_ceo= name_vi

    else:
            
        return JsonResponse({
            'result':False ,
            })

    draft.save()



    
    return JsonResponse({
        'result':True,
        })




#기안서 프린트
def draft_print(request,id=None):

    draft = Draft.objects.get(id = id)

    
    depart = COMMCODE.objects.get(id = id)
    type = COMMCODE.objects.get(commcode = draft.type ,commcode_grp='DRAFT_TYPE',upper_commcode='000007' )

    file_list= []
    file_list_query = Board_File.objects.filter(board_id=id, board_type='DRAFT').values('title').order_by('registered_date')


    file_rowspan = len(file_list_query)
    if file_rowspan == 0 :
        file_rowspan = 1
    else:
        tmp = 1
        for file in file_list_query:

            file_list.append({
                'id':tmp,
                'title':file['title'],
                })
            tmp +=1

    #diagnostic = reception.diagnosis.diagnosis
    return render(request,
    'Draft_Form/basic_form.html',
            {
                'type_vi':type.commcode_name_vi,
                'type_ko':type.commcode_name_ko,

                'depart_vi':depart.commcode_name_vi,
                'depart_ko':depart.commcode_name_ko,

                'date_registered':draft.date_registered[0:10],

                'status':draft.status,

                'title':draft.title,
                'contents':draft.contents,

                'file_rowspan':file_rowspan,
                'file_list':file_list,
                
            },
        )

    
    return JsonResponse({
        'result':True,
        })




#고객 관리
def customer_manage(request):
    

    return render(request,
        'Manage/CRM.html',
            {

            }
        )


def customer_manage_get_patient_list(request):

    category = request.POST.get('category')
    string = request.POST.get('string')


    argument_list = [] 
    if category=='':
        argument_list.append( Q(**{'name_kor__icontains':string} ) )
        argument_list.append( Q(**{'name_eng__icontains':string} ) )
        argument_list.append( Q(**{'id__icontains':string} ) ) 
        argument_list.append( Q(**{'phone__icontains':string} ) ) 
        argument_list.append( Q(**{'date_of_birth__icontains':string} ) ) 
    elif category=='name':
        argument_list.append( Q(**{'name_kor__icontains':string} ) )
        argument_list.append( Q(**{'name_eng__icontains':string} ) )
    elif category=='chart':
        argument_list.append( Q(**{'id__icontains':string} ) ) 
    elif category=='date_of_birth':
        argument_list.append( Q(**{'date_of_birth__icontains':string} ) ) 
    elif category=='phone':
        argument_list.append( Q(**{'phone__icontains':string} ) ) 



    patients = Patient.objects.filter( functools.reduce(operator.or_, argument_list) )

    datas=[]
    for patient in patients:

        visits = Reception.objects.filter(patient_id = patient.id).count()

        total_amount = Reception.objects.filter(patient_id = patient.id).prefetch_related('payment__paymentrecord_set').aggregate(total_price=Sum('payment__paymentrecord__paid'))

        data = {}
        data.update({
            'id':patient.id,
            'chart':patient.get_chart_no(),
            'name_kor':patient.name_kor,
            'name_eng':patient.name_eng,
            'gender':patient.get_gender_simple(),
            'date_of_birth':patient.date_of_birth.strftime('%Y-%m-%d'),
            'phonenumber':patient.phone,
            'age' : patient.get_age(),
            'address':patient.address,

            'memo':patient.memo,
            'date_registered':patient.date_registered.strftime('%Y-%m-%d'),

            'visits':visits,
            'paid_total':0 if total_amount['total_price'] == None else total_amount['total_price'],
            })
        datas.append(data)


    page_context = request.POST.get('context_in_page',10)
    page = request.POST.get('page',1)
    paginator = Paginator(datas, page_context)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)



    return JsonResponse({
        'result':True,
        'datas':list(paging_data),

        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),


        })




def customer_manage_get_patient_info(request):
    patient_id = request.POST.get('patient_id')

    context={}
    patient = Patient.objects.get(pk=int(patient_id))


    context.update({
        'id':patient.id,
        'chart':patient.get_chart_no(),
        'name_kor':patient.name_kor,
        'name_eng':patient.name_eng,
        'date_of_birth':patient.date_of_birth,
        'gender':patient.gender,
        'email':patient.email,
        'nationality':patient.nationality,
        'phone':patient.phone,
        'address':patient.address,
        'memo':patient.memo,
        })
    return JsonResponse(context)


def customer_manage_get_patient_visit(request):

    patient_id = request.POST.get('patient_id')

    context={}
    receptions = Reception.objects.filter(patient_id=int(patient_id)).exclude(progress='deleted').order_by('recorded_date')



    datas = []
    for reception in receptions:
        payment = Payment.objects.filter(reception_id = reception.id).aggregate(paid_sum=Sum('paymentrecord__paid'))



        datas.append({
            'reception_id':reception.id,
            'depart':reception.depart.name,
            'doctor':reception.doctor.name_short,
            'paid':'0' if payment['paid_sum'] is None else payment['paid_sum'],
            'date_visited':reception.recorded_date.strftime('%Y-%m-%d %H:%M'),
            })

    context.update({
        'datas':datas,
        })
    return JsonResponse(context)


def customer_manage_get_patient_visit_history(request):

    reception_id = request.POST.get('reception_id')

    reception = Reception.objects.get(id = reception_id)
    res = True
    data = {}
    try:
        diagnosis = Diagnosis.objects.get(reception_id = reception.id)

        exam_set = ExamManager.objects.filter(diagnosis_id = diagnosis.id)
        test_set = TestManager.objects.filter(diagnosis_id = diagnosis.id)
        precedure_set = PrecedureManager.objects.filter(diagnosis_id = diagnosis.id)
        medicine_set = MedicineManager.objects.filter(diagnosis_id = diagnosis.id)

        exams = []
        for data in exam_set:
            exam = {}
            exam.update({
                    'name':data.exam.name,
                })
            exams.append(exam)

        tests = []
        for data in test_set:
            test = {}
            test.update({
                    'name':data.test.name,
                    'amount':data.amount,
                    'days':data.days,
                    'memo':data.memo,
                })
            tests.append(test)

        precedures = []
        for data in precedure_set:
            precedure = {}
            precedure.update({
                    'name':data.precedure.name,
                    'amount':data.amount,
                    'days':data.days,
                    'memo':data.memo,
                })
            precedures.append(precedure)

        medicines = []
        for data in medicine_set:
            medicine = {}
            medicine.update({
                    'name':data.medicine.name,
                    'amount':data.amount,
                    'days':data.days,
                    'memo':data.memo,
                    'unit':data.medicine.unit,
                })
            medicines.append(medicine)

        data = {'date':diagnosis.recorded_date.strftime('%Y-%m-%d'),
        'day':diagnosis.recorded_date.strftime('%a'),
        'subjective':reception.chief_complaint,
        'objective':diagnosis.objective_data,
        'assessment':diagnosis.assessment,
        'plan':diagnosis.plan,
        'diagnosis':diagnosis.diagnosis,
        'ICD': diagnosis.ICD,
        'icd_code': diagnosis.ICD_code,
        'recommendation':diagnosis.recommendation,
                

        'doctor':reception.doctor.name_kor,

        'exams':exams,
        'tests':tests,
        'precedures':precedures,
        'medicines':medicines,
        'amount':'null',
        }

    except Diagnosis.DoesNotExist:
        res = False

    return JsonResponse({
        'result':res,

        'data':data,
    })



def customer_manage_get_patient_sms_info(request):
    patient_id = request.POST.get('patient_id')
        
    context={}
    patient = Patient.objects.get(pk=int(patient_id))

    context.update({
        'id':patient.id,
        'name_kor':patient.name_kor,
        'name_eng':patient.name_eng,
        'phone':patient.phone,
        })
    return JsonResponse(context)



def manage_employee(request):

    #직급
    list_rank = []
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        query_rank= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='RANK').annotate(code = F('commcode'),name = F('commcode_name_ko')).values('code','name')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        query_rank= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='RANK').annotate(code = F('commcode'),name = F('commcode_name_en')).values('code','name')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        query_rank= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='RANK').annotate(code = F('commcode'),name = F('commcode_name_vi')).values('code','name')

    for data in query_rank:
        list_rank.append({
            'id':data['code'],
            'name':data['name']
            })

    #부서 - KBL
    list_depart_kbl = []
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        query_depart= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_KBL').annotate(code = F('commcode'),name = F('commcode_name_ko')).values('code','name')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        query_depart= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_KBL').annotate(code = F('commcode'),name = F('commcode_name_en')).values('code','name')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        query_depart= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_KBL').annotate(code = F('commcode'),name = F('commcode_name_vi')).values('code','name')

    for data in query_depart:
        list_depart_kbl.append({
            'id':data['code'],
            'name':data['name']
            })

    #부서 - 병원
    list_depart_clinic = []
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        query_depart= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_CLICINC').annotate(code = F('commcode'),name = F('commcode_name_ko')).values('code','name','id')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        query_depart= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_CLICINC').annotate(code = F('commcode'),name = F('commcode_name_en')).values('code','name','id')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        query_depart= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_CLICINC').annotate(code = F('commcode'),name = F('commcode_name_vi')).values('code','name','id')

    for data in query_depart:
        if data['code'] == 'DOCTOR':
            temp_commcode = COMMCODE.objects.get(id = data['id'])
            data['code'] += '_' + temp_commcode.se1
        list_depart_clinic.append({
            'id':data['code'],
            'name':data['name']
            })



    #사원 구분 
    list_division = []
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        query_division= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='EMPLOYEE DIVISION').annotate(code = F('commcode'),name = F('commcode_name_ko')).values('code','name')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        query_division= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='EMPLOYEE DIVISION').annotate(code = F('commcode'),name = F('commcode_name_en')).values('code','name')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        query_division= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='EMPLOYEE DIVISION').annotate(code = F('commcode'),name = F('commcode_name_vi')).values('code','name')
    for data in query_division:
        list_division.append({
            'id':data['code'],
            'name':data['name']
            })

    #재직 상태
    list_status = []
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        query_status= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='EMPLOYEE STATUS').annotate(code = F('commcode'),name = F('commcode_name_ko')).values('code','name')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        query_status= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='EMPLOYEE STATUS').annotate(code = F('commcode'),name = F('commcode_name_en')).values('code','name')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        query_status= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='EMPLOYEE STATUS').annotate(code = F('commcode'),name = F('commcode_name_vi')).values('code','name')
    for data in query_status:
        list_status.append({
            'id':data['code'],
            'name':data['name']
            })

    return render(request,
        'Manage/employee_manage.html',
            {
                'list_rank':list_rank,
                'list_depart_clinic':list_depart_clinic,
                'list_depart_kbl':list_depart_kbl,
                'list_division':list_division,
                'list_status':list_status,
            }
        )


def employee_search(request):
    string = request.POST.get('string')
    division_type = request.POST.get('division_type')
    depart_filter = request.POST.get('depart_filter')


    kwargs = {}
    kwargs['is_active'] = True # 기본 
    if division_type is not '':
        kwargs['division_type'] = division_type
    if depart_filter is not '':
        if 'DOCTOR' in depart_filter:
            split_depart = depart_filter.split('_')

            kwargs['depart'] = 'DOCTOR'
            kwargs['depart_doctor'] = split_depart[1]
        else:
            kwargs['depart'] = depart_filter



    query_user = User.objects.filter(**kwargs)


    list_user = []
    for user in query_user:
       list_user.append({
               'id':user.id,
               'division_type':'' if user.division_type is None else user.division_type,
               'depart':'' if user.depart is None else user.depart,
               'rank':'' if user.rank is None else user.rank,
               'user_id':'' if user.user_id is None else user.user_id,
               'name_ko':'' if user.name_ko is None else user.name_ko,
               'name_en':'' if user.name_en is None else user.name_en,
               'name_vi':'' if user.name_vi is None else user.name_vi,
               'gender':'' if user.gender is None else user.gender,
               'date_of_birth':'' if user.date_of_birth is None else user.date_of_birth,
               'phone_number1':'' if user.phone_number1 is None else user.phone_number1,
               'phone_number2':'' if user.phone_number2 is None else user.phone_number2,
               'email':'' if user.email is None else user.email,
               'date_of_employment':'' if user.date_of_employment is None else user.date_of_employment,
               'status':'' if user.status is None else user.status,
               
           })
    
    page_context = request.POST.get('context_in_page',10)
    page = request.POST.get('page',1)
    paginator = Paginator(list_user, page_context)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)



    return JsonResponse({
        'result':True,
        'datas':list(paging_data),

        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),


        })




@login_required
def employee_check_id(request):
       
    user_id = request.POST.get('user_id')

    try:
        user = User.objects.get(user_id = user_id)
        result = False
    except User.DoesNotExist:
        result = True

    return JsonResponse({
        'result':result,
        })



@login_required
def employee_add_edit(request):

    id = request.POST.get('id')

    user_id = request.POST.get('user_id')
    password = request.POST.get('password')
    name_ko = request.POST.get('name_ko')
    name_en = request.POST.get('name_en')
    name_vi = request.POST.get('name_vi')
    gender = request.POST.get("gender")
    phone1 = request.POST.get('phone1')
    phone2 = request.POST.get('phone2')
    date_of_birth = request.POST.get('date_of_birth')
    email = request.POST.get('email')
    address = request.POST.get('address')

    rank = request.POST.get('rank')
    depart = request.POST.get('depart')
    division = request.POST.get('division')

    status = request.POST.get('status')
    date_of_employment = request.POST.get('date_of_employment')
    remark = request.POST.get('remark')

    
    try:
        user = User.objects.get(pk = id)
        user.lastest_modified_date = datetime.datetime.now()
    except User.DoesNotExist:
        user = User()
        user.set_password(password)
    

    user.user_id = user_id
    #기본 정보
    
    user.name_ko = name_ko
    user.name_en = name_en
    user.name_vi = name_vi
    user.gender = gender
    user.phone_number1 = phone1
    user.phone_number2 = phone2
    user.date_of_birth = date_of_birth
    user.email = email
    user.address = address

    user.status = status
    user.date_of_employment = date_of_employment
    user.memo = remark


    #권한 설정
    user.rank = rank
    user.depart = depart
    user.division_type = division

    #권한 설정 #superuser / staff / 
    user.save()
    if "DOCTOR" in depart: #의사 권한은 별도로 ... 
        str_split = depart.split('_')
        
        print(str_split)
        commcode = COMMCODE.objects.get(upper_commcode = '000002', commcode_grp='DEPART_CLICINC', commcode='DOCTOR',se1 = str_split[1])
       
        user.depart_doctor = str_split[1]
        user.depart = str_split[0]


        try:#의사 정보 변경
            doctor = Doctor.objects.get(user_id=user.id)
        except Doctor.DoesNotExist:#의사 신규
            doctor = Doctor()

        depart = Depart.objects.get(name = user.depart_doctor)
        doctor.depart = depart
        doctor.name_kor = name_en
        doctor.name_eng = name_en
        doctor.name_short = name_en
        doctor.user_id = user.id

        user.save()
        doctor.save()

    

    return JsonResponse({
        'result':True,
        })

@login_required
def employee_add_edit_get(request):
     
    id = request.POST.get('id')

    user = User.objects.get(pk = id)

    if user.depart == 'DOCTOR':
        depart = user.depart + "_" + user.depart_doctor


    else:
        depart = user.depart

    return JsonResponse({
        'result':True,

        "user_id" : user.user_id,
        "name_ko" : user.name_ko,
        "name_en" : user.name_en,
        "name_vi" : user.name_vi,
        "gender" : user.gender,
        "phone1" : user.phone_number1,
        "phone2" : user.phone_number2,
        "date_of_birth" : user.date_of_birth,
        "email" : user.email,
        "address" : user.address,

        "rank" : user.rank,
        "depart" : depart,
        "division" : user.division_type,

        "status" : user.status,
        "date_of_employment" : user.date_of_employment,
        "remark" : user.memo,

        })




@login_required
def employee_delete(request):

    id = request.POST.get('id')

    date_of_resignation = request.POST.get('date_of_resignation')
    resignation_reason = request.POST.get('resignation_reason')

    user = User.objects.get(pk = id)

    user.is_active = False

    user.status = 'RESIGNED'
    user.date_of_resignation = date_of_resignation
    user.resignation_memo = resignation_reason

    user.lastest_modified_date = datetime.datetime.now()

    user.save()

    return JsonResponse({
        'result':True,
        
        })


@login_required
def employee_change_password(request):
    id = request.POST.get('id')
    password = request.POST.get('password')

    user = User.objects.get(pk = id)
    user.set_password(password)
    user.save()

    return JsonResponse({
        'result':True,
        })


@login_required
def board_list(request,id=None):

    if id == None:
        id = request.POST.get('selected_content',None)
    
    current_language = request.session[translation.LANGUAGE_SESSION_KEY]
    if current_language == 'ko':
         fname = F('commcode_name_ko')
    elif current_language == 'en':
        fname = F('commcode_name_en')
    elif current_language == 'vi':
        fname = F('commcode_name_vi')
    #초기 - Division
    list_division = COMMCODE.objects.filter(upper_commcode = '000005',commcode_grp = 'BOARD_DIVISION')

    dict_division = {}
    for division in list_division.annotate(code = F('commcode'),name = fname).values('code','name'):
        dict_division.update({
            division['code'] : division['name']
            })



    #content
    content = None;
    #게시글 오픈 유무
    if id is not '':
        try:
            read_page = Board_Contents.objects.get(id = id )
            creator = User.objects.get(id = int(read_page.creator))
            #조회수 카운트
            ##자기 자신을 제외한 유저에만 해당
            if creator.id != request.user.id:
                date_min = datetime.datetime.combine(datetime.datetime.now().date(), datetime.time.min)
                date_max = datetime.datetime.combine(datetime.datetime.now().date(), datetime.time.max)

                is_new = Board_View_Log.objects.filter(registered_date__range = (date_min, date_max),user_id = request.user.id, board_id = read_page.id, ).count()
                if is_new == 0 : #오늘 안봤으면 추가
                    new_view = Board_View_Log()
                    new_view.board_id = read_page.id
                    new_view.user_id = request.user.id
                    new_view.save()

                    read_page.view_count += 1
                    read_page.save()

            #페이지 상세 정보 
            
            
            comments_count = Board_Comment.objects.filter(content_id = id, use_yn = 'Y').count()
            content = {
                'id':read_page.id,
                'title':read_page.title,
                'contents':read_page.contents,
                'creator':creator.user_id,
                'date':read_page.created_date.strftime('%Y-%m-%d %H:%M'),
                'views': read_page.view_count,
                'comments_count':comments_count,
                }

            query_file = Board_File.objects.filter(board_id = id,board_type='BASIC')

            list_file = []
            for file in query_file:
                list_file.append({
                    'id':file.id,
                    'file_name':file.file.url,
                    'origin_name':file.origin_name,
                    })
            content.update({
                'list_file': list_file,
                'list_file_count':query_file.count(),
                })

        except Board_Contents.DoesNotExist:
            content = None;



   

    #search filter
    kwargs = {}
    kwargs['use_yn'] = 'Y' # 기본 
    kwargs['board_type'] = 'BASIC' #일반 게시판

    kwargs['is_KBL'] = 'N'
    if request.META['SERVER_PORT'] == '8888':#경천애인
        kwargs['is_KBL'] = 'Y'


    search_string = request.POST.get('search_string','')
    view_division_filter = request.POST.get('view_division_filter','')
    if view_division_filter != '':
        kwargs['options'] = view_division_filter

    users = User.objects.all()
    if search_string == '':
        query = Board_Contents.objects.filter(**kwargs).order_by('-top_seq','-created_date')

    else:
        argument_list = [] 

        argument_list.append( Q(**{'title__icontains':search_string} ) ) 
        argument_list.append( Q(**{'contents__icontains':search_string} ) ) 

        #list_search_user = users.filter()
        query = Board_Contents.objects.filter(functools.reduce(operator.or_, argument_list),**kwargs).order_by('-created_date')


    #목록 정렬
    contents_list = []
    for item in query:
        comment_count = Board_Comment.objects.filter(content_id = item.id, use_yn = 'Y').count()
        creator = users.get(id = int(item.creator))
        file_count = Board_File.objects.filter(board_id = item.id).count()


        contents_list.append({
            'id':item.id,
            'is_notice':item.top_seq,
            'division':dict_division[item.options],
            'title':item.title,
            'creator':creator.user_id,
            'date':item.created_date.strftime('%Y-%m-%d %H:%M'),
            'comment_count':comment_count,
            'is_file':False if file_count is 0 else True,
            'view_count':item.view_count,
            })


    #페이지네이션
    page = request.POST.get('page',1)
    view_contents_count = request.POST.get('view_contents_count',10);
    paginator = Paginator(contents_list, view_contents_count)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)

    return render(request,
        'board/list.html',
            {
                'content_count':query.count(),
                'content':content,
                #'contents_list':contents_list,
                'contents_list':paging_data,

                'search_string':search_string,
                'view_division_filter':view_division_filter,
                'view_contents_count':view_contents_count,
                'dict_division':dict_division,

                'page':int(page),
                'page_range':list( range(paging_data.paginator.page_range.start, paging_data.paginator.page_range.stop) ) ,
                'page_range_start':paging_data.paginator.page_range.start,
                'page_range_stop':paging_data.paginator.page_range.stop,
                'page_number':paging_data.number,
                'has_previous':paging_data.has_previous(),
                'has_next':paging_data.has_next(),
            }
        )



@login_required
def board_create_edit(request,id=None):
    
    division_selected= None
    option_err=''
    is_top= '0'
    #language
    lang = ''
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        lang = 'vi-VN'
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        lang = 'ko-KR'
    else:
        lang = 'en-US'
    
    list_file = []
    if id is None: # 새글
        load_contents = Board_Contents()
        form = board_form()

        file_form = board_file_form()

        load_contents.creator = request.user.id
        if request.META['SERVER_PORT'] == '8888':#경천애인
            load_contents.is_KBL = 'Y' 
    else: # 글 수정 
        load_contents = Board_Contents.objects.get(id = id)
        form = board_form(instance = load_contents)
        division_selected = load_contents.options

        is_top = load_contents.top_seq

        file_form = board_file_form()
        query_file = Board_File.objects.filter(board_id = id,board_type='BASIC')
        for file in query_file:
            list_file.append({
                'id':file.id,
                'file_name':file.file.url,
                'origin_name':file.origin_name,
                })


    #저장
    if request.method == 'POST':
        form = board_form(request.POST)
        file_form = board_file_form(request.POST, request.FILES)   
        files = request.FILES.getlist('file') 


        select_valid = False
        division_selected = request.POST.get('select_division',None)
        if division_selected is not None:
            select_valid = True


        if form.is_valid() and file_form.is_valid() and select_valid:
            load_contents.board_type = 'BASIC'
            load_contents.title = form.cleaned_data['title']
            load_contents.contents = form.cleaned_data['contents']
            load_contents.options = division_selected

            is_notice = request.POST.get('top_seq','off')
            if is_notice == 'on':
                load_contents.top_seq = 1
   

            
            load_contents.lastest_modifier = request.user.id
            load_contents.lastest_modified_date = datetime.datetime.now()

            load_contents.save()

            #file save
            for f in files:
                file_instance = Board_File(file=f, board_id = load_contents.pk)
                file_instance.origin_name = f.name
                file_instance.board_type = 'BASIC'
                file_instance.user = request.user.user_id

                file_instance.save()
 

            if id is None:
                return HttpResponseRedirect('./../' + str(load_contents.pk))
            else:
                return HttpResponseRedirect('./../../' + str(load_contents.pk))
        else:
            pass#form = board_form(instance=profile)




    
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    list_division = []
    query_division= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='BOARD_DIVISION').annotate(code = F('commcode'),name = f_name ).values('code','name')
    for data in query_division:
        list_division.append({
            'id':data['code'],
            'name':data['name']
            })
    

    return render(request,
        'board/create_edit.html',
            {
                'form':form,
                'file_form':file_form,
                'lang':lang,

                'list_file': list_file if list_file else None,
                'list_file_count':len(list_file),

                'list_division':list_division,
                'division_selected':division_selected, #division
                'option_err':option_err,
                'is_top':is_top
            }
        )



@login_required
def board_delete(request,id):

    try:
        content = Board_Contents.objects.get(id = id)
        content.use_yn = 'N'
        content.save()

    except Board_Contents.DoesNotExist:
        pass


    return HttpResponseRedirect('./../')


@login_required
def board_delete_file(request):
    id = request.POST.get('id')

    file = Board_File.objects.get(pk = id)


    file.delete()


    return JsonResponse({
        'result':True,
        })



@login_required
def board_comment_get(request):
    content_id = request.POST.get('content_id')

    #set user data
    user_dict = {}
    users = User.objects.all().values('id','user_id')
    for user in users:
        user_dict[user['id']] = user['user_id']

    list_comment = []
    query = Board_Comment.objects.filter(content_id = content_id,use_yn = 'Y').order_by('orderno').values()

    
    for data in query:
        comment = {}
        comment.update({
            'id':data['id'],
            'user_id':data['creator'],
            'user':user_dict[ int(data['creator']) ],
            'comment':data['comment'],
            'datetime':data['created_date'].strftime('%Y-%m-%d %H:%M'),
            'depth':data['depth'],
            })

        if request.user.id is int(data['creator']):
            comment.update({'is_creator':True})
        else:
            comment.update({'is_creator':False})

        list_comment.append(comment)




    return JsonResponse({
        'result':True,
        'list_comment':list_comment,
        'conntents_count':query.count(),
        })


def board_comment_set_seq(input_data):
    



    return

def board_comment_add(request,comment_id=None):
    content_id = request.POST.get('content_id')
    comment = request.POST.get('comment',None)
    upper_id = request.POST.get('upper_id',None)


    new_comment = Board_Comment()




    if upper_id != '':#대댓글
        upper = Board_Comment.objects.get(pk = upper_id)
        new_comment.depth = upper.depth + 1
        
        try:
            check = Board_Comment.objects.filter(content_id = content_id, orderno__gt = upper.orderno, depth__lte = upper.depth,).order_by('orderno')[:1]
            
            new_comment.orderno = check[0].orderno
        except IndexError:
            check = Board_Comment.objects.filter(content_id = content_id, ).order_by('-orderno')[:1]
           
            new_comment.orderno = check[0].orderno + 1
        

        #저장 전 sequence 자리 비우기
        query_set = Board_Comment.objects.filter(content_id = content_id, orderno__gte = new_comment.orderno ,).order_by('orderno')
        
        for query in query_set:
            query.orderno += 1
            
            query.save()

          
    else:#첫번째 뎁스 댓글
        check = Board_Comment.objects.filter(content_id = content_id, ).order_by('orderno')
        
        if check.count() != 0:
            check_last = check.last()
            new_comment.orderno = check_last.orderno + 1




    new_comment.content_id = content_id
    new_comment.comment = comment
    new_comment.creator = request.user.id


    new_comment.save()

    return JsonResponse({
        'result':True,
        })






def board_comment_edit(request,id):
    content_content = request.POST.get('comment')
    comment = Board_Comment.objects.get(id = id)
    comment.comment = content_content
    comment.last_modified_date = datetime.datetime.now()
    comment.save()

    return JsonResponse({
        'result':True,
        })


def board_comment_delete(request,id):#Comment ID
    content_id = request.POST.get('content_id')

    comment = Board_Comment.objects.get(id=id)
    comment.use_yn = 'N'
    comment.save()


    return JsonResponse({
        'result':True,
        })








@login_required
def board_work_list(request,id=None):
    def_date = '0000-00-00 00:00:00'


    if id == None:
        id = request.POST.get('selected_content',None)
    
    current_language = request.session[translation.LANGUAGE_SESSION_KEY]
    if current_language == 'ko':
         fname = F('commcode_name_ko')
    elif current_language == 'en':
        fname = F('commcode_name_en')
    elif current_language == 'vi':
        fname = F('commcode_name_vi')

    #set user data
    user_dict = {}
    users = User.objects.all().values('id','user_id')
    for user in users:
        user_dict.update({
            user['id'] : user['user_id']
            })


    #초기 - Division
    list_division = COMMCODE.objects.filter(upper_commcode = '000005',commcode_grp = 'BOARD_WORK_DIVISION')

    dict_division = {}
    for division in list_division.annotate(code = F('commcode'),name = fname).values('code','name'):
        dict_division.update({
            division['code'] : division['name']
            })

    # - Depart
    dict_depart = {}
    if request.META['SERVER_PORT'] == '8888':#경천애인
        ##경천
        query_depart_kbl= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_KBL',upper_commcode ='000002' )
        for data in query_depart_kbl.annotate(code = F('commcode'),name = fname).values('code','name'):
            dict_depart.update({
                data['code'] : data['name']
                })
    else:
        ##IMEDI
        query_depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_CLICINC',upper_commcode ='000002' )
        for data in query_depart_medical.annotate(code = F('commcode'),name = fname).values('code','name'):
            dict_depart.update({
                data['code'] : data['name']
                })

    ##관리자
    query_depart_admin= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_ADMIN',upper_commcode ='000002' )
    for data in query_depart_admin.annotate(code = F('commcode'),name = fname).values('code','name'):
        dict_depart.update({
            data['code'] : data['name']
            })

    # - Status
    dict_status ={}
    query_status= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='BOARD_STATUS',upper_commcode ='000005' )
    for data in query_status.annotate(code = F('commcode'),name = fname).values('code','name'):
        dict_status.update({
            data['code'] : data['name']
            })

    #content
    content = None;
    #게시글 오픈 유무
    if id is not '':
        try:
            read_page = Board_Contents.objects.get(id = id )
            creator = User.objects.get(id = int(read_page.creator))
            #조회수 카운트
            ##자기 자신을 제외한 유저에만 해당
            if creator.id != request.user.id:
                #date_min = datetime.datetime.combine(datetime.datetime.now().date(), datetime.time.min)
                #date_max = datetime.datetime.combine(datetime.datetime.now().date(), datetime.time.max)

                is_new = Board_View_Log.objects.filter(user_id = request.user.id, board_id = read_page.id).count()
                if is_new == 0 : #오늘 안봤으면 추가
                    new_view = Board_View_Log()
                    new_view.board_id = read_page.id
                    new_view.user_id = request.user.id
                    new_view.save()

                    read_page.view_count += 1
                    read_page.save()

            #페이지 상세 정보 
            
            
            comments_count = Board_Comment.objects.filter(content_id = id, use_yn = 'Y').count()
            content = {
                'id':read_page.id,
                'title':read_page.title,
                'contents':read_page.contents,
                'creator':creator.user_id,
                'date':read_page.created_date.strftime('%Y-%m-%d %H:%M'),
                'views': read_page.view_count,
                'comments_count':comments_count,
                'expected_date':'' if read_page.date_to_be_done == def_date else read_page.date_to_be_done,
                'due_date':'' if read_page.date_done == def_date else read_page.date_done,
                }
            selected_status = read_page.status
            query_file = Board_File.objects.filter(board_id = id,board_type='BASIC')

            list_file = []
            for file in query_file:
                list_file.append({
                    'id':file.id,
                    'file_name':file.file.url,
                    'origin_name':file.origin_name,
                    })
            content.update({
                'list_file': list_file,
                'list_file_count':query_file.count(),
                'selected_status':selected_status,
                })

            



        except Board_Contents.DoesNotExist:
            content = None;


    #search filter
    argument_list = [] 
    kwargs = {}
    kwargs['use_yn'] = 'Y' # 기본 
    kwargs['board_type'] = 'COWORK' #협업 게시판

    search_string = request.POST.get('search_string','')
    view_division_filter = request.POST.get('view_division_filter','')
    if view_division_filter != '':
        kwargs['options'] = view_division_filter

    kwargs['is_KBL'] = 'N'
    if request.META['SERVER_PORT'] == '8888':#경천애인
        kwargs['is_KBL'] = 'Y'


    if request.user.depart != 'ADMIN':
        argument_list.append( Q(**{'depart_to':request.user.depart} ) ) 
        argument_list.append( Q(**{'depart_from':request.user.depart} ) ) 

    if search_string != '':
        argument_list.append( Q(**{'title__icontains':search_string} ) ) 
        argument_list.append( Q(**{'contents__icontains':search_string} ) ) 
    
    if request.user.depart == 'ADMIN' and search_string == '':
        query = Board_Contents.objects.filter(**kwargs).order_by('-created_date')
    else:
        query = Board_Contents.objects.filter(functools.reduce(operator.or_, argument_list),**kwargs).order_by('-created_date')
    users = User.objects.all()

    #목록 정렬
    contents_list = []
    for item in query:
        comment_count = Board_Comment.objects.filter(content_id = item.id, use_yn = 'Y').count()
        creator = users.get(id = int(item.creator))
        file_count = Board_File.objects.filter(board_id = item.id).count()


        #댓글 불러오기
        list_comment = []
        query = Board_Comment.objects.filter(content_id = item.id,use_yn = 'Y').order_by('orderno').values()
        no=0
        for data in query:
            comment = {}
            comment.update({
                'no':no,
                'id':data['id'],
                'user_id':data['creator'],
                'user':user_dict[ int(data['creator']) ],
                'comment':data['comment'],
                'datetime':data['created_date'].strftime('%Y-%m-%d %H:%M'),
                'depth':data['depth'],
                'orderno':data['orderno'],
                })
            no +=1

            if request.user.id is int(data['creator']):
                comment.update({'is_creator':True})
            else:
                comment.update({'is_creator':False})

            list_comment.append(comment)


        
        contents_list.append({
            'id':item.id,
            'is_notice':item.top_seq,
            'division':dict_division[item.options],
            'title':item.title,
            'creator':creator.user_id,
            'date':item.created_date.strftime('%Y-%m-%d'),
            'depart_from':dict_depart[item.depart_from],
            'depart_to':dict_depart[item.depart_to],
            'status':dict_status[item.status],
            'is_file':False if file_count is 0 else True,
            'view_count':item.view_count,
            
            'list_comment':list_comment,
            'expected_date':'' if item.date_to_be_done == def_date else item.date_to_be_done,
            'due_date':'' if item.date_done == def_date else item.date_done,

            #'comment_count':comment_count,
            })




    #페이지네이션
    page = request.POST.get('page',1)
    view_contents_count = request.POST.get('view_contents_count',10);
    paginator = Paginator(contents_list, view_contents_count)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)

    return render(request,
        'board_work/list.html',
            {
                'content_count':query.count(),
                'content':content,
                #'contents_list':contents_list,
                'contents_list':paging_data,

                'search_string':search_string,
                'view_division_filter':view_division_filter,
                'view_contents_count':view_contents_count,
                'dict_division':dict_division,
                
                
                'dict_status':dict_status,


                'page':int(page),
                'page_range':list( range(paging_data.paginator.page_range.start, paging_data.paginator.page_range.stop) ) ,
                'page_range_start':paging_data.paginator.page_range.start,
                'page_range_stop':paging_data.paginator.page_range.stop,
                'page_number':paging_data.number,
                'has_previous':paging_data.has_previous(),
                'has_next':paging_data.has_next(),
            }
        )



@login_required
def board_work_create_edit(request,id=None):
    option_err=''
    is_top= '0'
    division_selected = None
    depart_to_selected = None
    #language
    
    list_file = []
    if id is None: # 새글
        load_contents = Board_Contents()
        form = board_form()

        file_form = board_file_form()

        load_contents.creator = request.user.id
        load_contents.depart_from = request.user.depart
        if request.user.depart =='DOCTOR':
            load_contents.depart_from = request.user.depart + "_" + request.user.depart_doctor

            
        if request.META['SERVER_PORT'] == '8888':#경천애인
            load_contents.is_KBL = 'Y'
    else: # 글 수정 
        load_contents = Board_Contents.objects.get(id = id)
        form = board_form(instance = load_contents)
        division_selected = load_contents.options

        is_top = load_contents.top_seq

        file_form = board_file_form()
        query_file = Board_File.objects.filter(board_id = id)
        for file in query_file:
            list_file.append({
                'id':file.id,
                'file_name':file.file.url,
                'origin_name':file.origin_name,
                })


    #저장
    if request.method == 'POST':
        form = board_form(request.POST)
        file_form = board_file_form(request.POST, request.FILES)   
        files = request.FILES.getlist('file') 

        division_select_valid = False
        division_selected = request.POST.get('select_division',None)
        if division_selected is not None:
            division_select_valid = True

        depart_to_select_valid = False
        depart_to_selected = request.POST.get('select_depart_to',None)
        if depart_to_selected is not None:
            depart_to_select_valid = True

        satus_selected = request.POST.get('select_status','SUBMIT')

        if form.is_valid() and file_form.is_valid() and division_select_valid and depart_to_select_valid:
            load_contents.board_type = 'COWORK'
            load_contents.title = form.cleaned_data['title']
            load_contents.contents = form.cleaned_data['contents']
            load_contents.options = division_selected
            load_contents.status = satus_selected
            load_contents.depart_to = depart_to_selected


            load_contents.lastest_modifier = request.user.id
            load_contents.lastest_modified_date = datetime.datetime.now()
            
            load_contents.save()

            #file save
            for f in files:
                file_instance = Board_File(file=f, board_id = load_contents.pk)
                file_instance.origin_name = f.name
                file_instance.board_type = 'BASIC'
                file_instance.user = request.user.user_id
                file_instance.save()
 

            if id is None:
                return HttpResponseRedirect('./../' + str(load_contents.pk))
            else:
                return HttpResponseRedirect('./../../' + str(load_contents.pk))
        else:
            pass#form = board_form(instance=profile)




    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    
    #구분
    list_division = []
    query_division= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='BOARD_WORK_DIVISION').annotate(code = F('commcode'),name = f_name ).values('code','name')
    for data in query_division:
        list_division.append({
            'id':data['code'],
            'name':data['name']
            })
    #부서

    ##경천
    list_depart = []
    if request.META['SERVER_PORT'] == '8888':#경천애인
        query_depart_kbl= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_KBL',upper_commcode ='000002' ).annotate(code = F('commcode'),name = f_name ).values('code','name')
        for data in query_depart_kbl:
            list_depart.append({
                'id':data['code'],
                'name':data['name']
                })
    else:
        ##IMEDI
        list_depart_medical = []
        query_depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('commcode'),name = f_name ).values('code','name')
        for data in query_depart_medical:
            list_depart.append({
                'id':data['code'],
                'name':data['name']
                })

    #게시판 상태
    list_status = []
    query_status= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='BOARD_STATUS',upper_commcode ='000005' ).annotate(code = F('commcode'),name = f_name ).values('code','name')
    for data in query_status:
        list_status.append({
            'id':data['code'],
            'name':data['name']
            })


    return render(request,
        'board_work/create_edit.html',
            {
                'form':form,
                'file_form':file_form,

                'list_file': list_file if list_file else None,
                'list_file_count':len(list_file),

                'list_division':list_division,
                'division_selected':division_selected, #division
                'depart_to_selected':depart_to_selected,
                'list_depart':list_depart,
                'list_status':list_status,
                
                'option_err':option_err,

            }
        )


def board_work_comment_add(request):
    content_id = request.POST.get('content_id')
    comment = request.POST.get('comment',None)
    upper_id = request.POST.get('upper_id',None)


    new_comment = Board_Comment()




    if upper_id != '':#대댓글
        upper = Board_Comment.objects.get(pk = upper_id)
        new_comment.depth = upper.depth + 1
        
        try:
            check = Board_Comment.objects.filter(content_id = content_id, orderno__gt = upper.orderno, depth__lte = upper.depth,).order_by('orderno')[:1]
            
            new_comment.orderno = check[0].orderno
        except IndexError:
            check = Board_Comment.objects.filter(content_id = content_id, ).order_by('-orderno')[:1]
           
            new_comment.orderno = check[0].orderno + 1
        

        #저장 전 sequence 자리 비우기
        query_set = Board_Comment.objects.filter(content_id = content_id, orderno__gte = new_comment.orderno ,).order_by('orderno')
        
        for query in query_set:
            query.orderno += 1
            
            query.save()

          
    else:#첫번째 뎁스 댓글
        check = Board_Comment.objects.filter(content_id = content_id, ).order_by('orderno')
        
        def_date = '0000-00-00 00:00:00'

        expected_date = request.POST.get('expected_date',def_date)
        due_date = request.POST.get('due_date',def_date)
        select_status = request.POST.get('select_status')

        content = Board_Contents.objects.get(id = content_id)
        content.status = select_status
        content.date_to_be_done = expected_date
        content.date_done = due_date
        content.save()

        if check.count() != 0: # 완전 처음이 아닐때 
            check_last = check.last()
            new_comment.orderno = check_last.orderno + 1




    new_comment.content_id = content_id
    new_comment.comment = comment
    new_comment.creator = request.user.id


    new_comment.save()

    return JsonResponse({
        'result':True,
        })



def sms_send_sms(request):

    type = request.POST.get('type','')
    company = request.POST.get('company','')
    receiver = request.POST.get('receiver','')

    phone = request.POST.get('phone','')
    contents = request.POST.get('contents','')

    sms_send = sms_history()

    sms_send.type = type
    sms_send.company = company
    sms_send.receiver = receiver

    sms_send.sender = request.user.user_id
    sms_send.phone = phone
    sms_send.contents = contents
    sms_send.date_of_registered = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sms_send.save()

    return JsonResponse({
        'res':True,
        'id':sms_send.id,
        })


def sms_recv_result(request):

    context = {}
    msg_id = request.POST.get('msg_id',None)
    if msg_id is not None:
        result = request.POST.get('result','')
        code = request.POST.get('code','')


        res_sms = sms_history.objects.get(id = msg_id)


        res_sms.status = result
        res_sms.res_code = code
        res_sms.date_of_recieved = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        res_sms.save()


    return JsonResponse({
        'res':True,
        })


def sms_history_index(request):


    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')

    list_type = []
    query_division= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='SMS_TYPE',upper_commcode ='000008').annotate(code = F('commcode'),name = f_name ).values('code','name')
    for data in query_division:
        list_type.append({
            'id':data['code'],
            'name':data['name']
            })

    list_search = []
    query_division= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='SMS_SEARCH',upper_commcode ='000008').annotate(code = F('commcode'),name = f_name ).values('code','name')
    for data in query_division:
        list_search.append({
            'id':data['code'],
            'name':data['name']
            })


    return render(request,
    'Manage/sms_history.html',
        {
            'list_type':list_type,
            'list_search':list_search,
        }
    )


def sms_history_search(request):

    start = request.POST.get('start')
    end = request.POST.get('end')

    type = request.POST.get('type','')
    option = request.POST.get('option','')
    string = request.POST.get('string','')

    print(type)
    kwargs = {}
    if type != '':
        kwargs['type'] = type
    #if option != '':
    #    kwargs['type'] == type
    #if type != '':
    #    kwargs['type'] == type
    kwargs['is_KBL'] = 'N'
    if request.META['SERVER_PORT'] == '8888':#경천애인
        kwargs['is_KBL'] = 'Y'

    argument_list = [] 
    if string !='':
        if option != 'COMPANY':
            argument_list.append( Q(**{'receiver__icontains':string} ) )
        if option != 'PERSONAL':
            argument_list.append( Q(**{'company_name__icontains':string} ) )
        

    date_min = datetime.datetime.combine(datetime.datetime.strptime(start,"%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(end,"%Y-%m-%d").date(), datetime.time.max)


    list_sms_history = []

    if len(argument_list) != 0:
        query_sms = sms_history.objects.filter(functools.reduce(operator.or_, argument_list) ,**kwargs, date_of_registered__range = (date_min, date_max))
    else:
        query_sms = sms_history.objects.filter(**kwargs, date_of_registered__range = (date_min, date_max))

    for query in query_sms:
        list_sms_history.append({
            'id':query.id,
            'type':query.type,
            'company':query.company_name,
            'receiver':query.receiver,
            'phone':query.phone,
            'datetime':query.date_of_registered[0:16],
            'contents':query.contents,
            'sender':query.sender,
            'status':query.status,
            'remark':query.res_code,
            })


    #페이지네이션
    page = request.POST.get('page',1)
    view_contents_count = request.POST.get('context_in_page',10);

    paginator = Paginator(list_sms_history, view_contents_count)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)

    return JsonResponse({
        'res':True,

        'datas':list_sms_history, 

        'page':int(page),
        'page_range':list( range(paging_data.paginator.page_range.start, paging_data.paginator.page_range.stop) ) ,
        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),
        })


def sms_history_get(request):

    history_id = request.POST.get('id')
         
    context={}
    history = sms_history.objects.get(pk=int(history_id))

    context.update({
        'name':history.receiver,
        'phone':history.phone,
        'contents':history.contents,
        })

    return JsonResponse(context)



def statistics_test(request):


    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })

    
    
    return render(request,
    'statistics/statistics_test.html',
        {
            'depart_medical':depart_medical,
        }
    )


def statistics_procedure(request):


    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })

    
    
    return render(request,
    'statistics/statistics_procedure.html',
        {
            'depart_medical':depart_medical,
        }
    )

def statistics_medicine(request):
    
    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })

    
    
    return render(request, 
    'statistics/statistics_medicine.html',
        {
            'depart_medical':depart_medical,
        }
    )



def statistics_depart(request):
    
    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })

    
    
    return render(request, 
    'statistics/statistics_depart.html',
        {
            'depart_medical':depart_medical,
        }
    )


def statistics_search(request):
    type = request.POST.get('type')

    start = request.POST.get('start')
    end = request.POST.get('end')
    depart = request.POST.get('depart')

    kwargs = {}
    if depart!='':
        kwargs['depart'] = depart # 기본 


    date_min = datetime.datetime.combine(datetime.datetime.strptime(start,"%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(end,"%Y-%m-%d").date(), datetime.time.max)


    data_list = []

 
    total_revenue = 0
    


    if type == 'TEST':
        tests = Test.objects.all().order_by('code')
        for test in tests:
            price_sum = 0

            sub_query = Reception.objects.filter(
                **kwargs ,
                recorded_date__range = (date_min, date_max), 
                diagnosis__testmanager__test = test.id,
            ).exclude(
                progress='deleted'
                ).prefetch_related(
                'diagnosis__testmanager_set',
            )

            if not sub_query:
                continue

            for data in sub_query:
                price_sum = test.get_price(data.recorded_date)


            data_list.append({
                'id':test.id,
                'name':test.name,
                'name_vi':test.name_vie,
                'count':sub_query.count(),
                'price_sum':price_sum,
                })

            total_revenue += price_sum
    elif type == 'PROCEDURE':
        procedures = Precedure.objects.all().order_by('code')
        for procedure in procedures:
            price_sum = 0
            count = 0

            sub_query = Reception.objects.filter(
                **kwargs ,
                recorded_date__range = (date_min, date_max), 
                diagnosis__preceduremanager__precedure= procedure.id,
            ).exclude(
                progress='deleted'
                ).prefetch_related(
                'diagnosis__preceduremanager_set',
            )
            if not sub_query:
                continue

            for data in sub_query:
                preceduremanager_set = PrecedureManager.objects.filter(diagnosis_id = data.diagnosis.id,precedure_id = procedure.id)
                #price_sum = procedure.get_price(data.recorded_date)
                if preceduremanager_set.count() != 0:
                    for set_data in preceduremanager_set:
                        count += set_data.amount
                        price_sum += procedure.get_price(data.recorded_date) * set_data.amount


            data_list.append({
                'id':procedure.id,
                'name':procedure.name,
                'name_vi':procedure.name_vie,
                'count':count,
                'price_sum':price_sum,
                })

            total_revenue += price_sum

    elif type =='MEDICINE':

        medicines= Medicine.objects.all().order_by('code')
        for medicine in medicines:
            price_sum = 0
            count = 0

            sub_query = Reception.objects.filter(
                **kwargs ,
                recorded_date__range = (date_min, date_max), 
                diagnosis__medicinemanager__medicine= medicine.id,
            ).exclude(
                progress='deleted'
                ).prefetch_related(
                'diagnosis__medicinemanager_set',
            )

            if not sub_query:
                continue

            for data in sub_query: 
                print(data.id)
                medicine_set = MedicineManager.objects.filter(diagnosis_id = data.diagnosis.id, medicine_id=medicine.id)
                for set_data in medicine_set:
                    count += set_data.amount * set_data.days
                    price_sum += medicine.get_price(data.recorded_date) * set_data.amount * set_data.days
                    
            data_list.append({
                'id':medicine.id,
                'name':medicine.name,
                'name_vi':medicine.name_vie,
                'count':count,
                'price_sum':price_sum,
                })

            total_revenue += price_sum


    elif type == 'DEPART':
        departs = Depart.objects.all()
        
        for depart in departs:
            
            views_count = Reception.objects.filter(
                **kwargs ,
                recorded_date__range = (date_min, date_max), 
                depart_id = depart.id 
                ).exclude(progress='deleted').count()


            revenue_total = Reception.objects.filter(
                **kwargs ,
                recorded_date__range = (date_min, date_max), 
                depart_id = depart.id 
                ).exclude(
                progress='deleted'
                ).prefetch_related('payment__paymentrecord_set').aggregate(total_price=Sum('payment__paymentrecord__paid'))
                 
            data_list.append({
                'id':depart.id,
                'name':depart.name,
                'name_vi':depart.full_name_vie,
                'count':views_count,
                'price_sum':0 if revenue_total['total_price'] is None else revenue_total['total_price'],
                })




    return JsonResponse({
        'result':True,
        'datas':data_list,

        'total_revenue':total_revenue
        })









def statistics_customer_info(request):
    
    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })

    
    
    return render(request, 
    'statistics/statistics_customer_info.html',
        {
            'depart_medical':depart_medical,
        }
    )


def search_customer_info(request):

    start = request.POST.get('start')
    end = request.POST.get('end')
    depart = request.POST.get('depart')

    kwargs = {}
    if depart!='':
        kwargs['depart'] = depart # 기본 


    date_min = datetime.datetime.combine(datetime.datetime.strptime(start,"%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(end,"%Y-%m-%d").date(), datetime.time.max)



    data_list_gender = []
    data_list_nation = []
    data_list_payment_method = []
    data_list_age = []

    #gender
    gender_count = Reception.objects.filter(
                **kwargs ,
                recorded_date__range = (date_min, date_max), 
            ).exclude(
                progress='deleted'
            ).select_related(
                'patient'
            ).prefetch_related(
                'payment__paymentrecord_set'
            ).values(
                'patient__gender'
            ).annotate(
                gender_count = Count('patient__gender'),
                total_price = Sum('payment__paymentrecord__paid')
            )

    for data in gender_count:
        data_list_gender.append({
            'name':data['patient__gender'],
            'count':data['gender_count'],
            'price_sum':data['total_price'],
            })

    #Nationality
    nation_count = Reception.objects.filter(
                **kwargs ,
                recorded_date__range = (date_min, date_max), 
            ).exclude(
                progress='deleted'
            ).select_related(
                'patient'
            ).prefetch_related(
                'payment__paymentrecord_set'
            ).values(
                'patient__nationality'
            ).annotate(
                nation_count= Count('patient__nationality'),
                total_price = Sum('payment__paymentrecord__paid')
            )

    for data in nation_count:
        data_list_nation.append({
            'name':data['patient__nationality'],
            'count':data['nation_count'],
            'price_sum':data['total_price'],
            })

    #payment_method
    payment_method_count = Reception.objects.filter(
                **kwargs ,
                recorded_date__range = (date_min, date_max), 
            ).exclude(
                progress='deleted'
            ).prefetch_related(
                'payment__paymentrecord_set'
            ).values(
                'payment__paymentrecord__method'
            ).annotate(
                method_count= Count('payment__paymentrecord__method'),
                total_price = Sum('payment__paymentrecord__paid')
            )

    for data in payment_method_count:
        if data['payment__paymentrecord__method'] == None:
            continue

        data_list_payment_method.append({
            'name':data['payment__paymentrecord__method'],
            'count':data['method_count'],
            'price_sum':0 if data['total_price'] is None else data['total_price'],
            })



 #age
    today = datetime.datetime.now()

   
    age_list=[
            {#0~9세
                'name':'0~9',
                'date_start':today,
                'date_end': (today - datetime.timedelta(days=365.25 * 10 ) ).date(),
            },
            {#10~19세
                'name':'10~19',
                'date_start':(today - datetime.timedelta(days=365.25 * 10 ) ).date(),
                'date_end': (today - datetime.timedelta(days=365.25 * 20 ) ).date(),
            },
            {#20~29세
                'name':'20~29',
                'date_start':(today - datetime.timedelta(days=365.25 * 20 ) ).date(),
                'date_end': (today - datetime.timedelta(days=365.25 * 30 ) ).date(),
            },
            {#30~39세
                'name':'30~39',
                'date_start':(today - datetime.timedelta(days=365.25 * 30 ) ).date(),
                'date_end': (today - datetime.timedelta(days=365.25 * 40 ) ).date(),
            },
            {#40~49세
                'name':'40~49',
                'date_start':(today - datetime.timedelta(days=365.25 * 40 ) ).date(),
                'date_end': (today - datetime.timedelta(days=365.25 * 50 ) ).date(),
            },
            {#50~59세
                'name':'50~59',
                'date_start':(today - datetime.timedelta(days=365.25 * 50 ) ).date(),
                'date_end': (today - datetime.timedelta(days=365.25 * 60 ) ).date(),
            },
            {#60~69세
                'name':'60~69',
                'date_start':(today - datetime.timedelta(days=365.25 * 60 ) ).date(),
                'date_end': (today - datetime.timedelta(days=365.25 * 70 ) ).date(),
            },
            {#70~79세
                'name':'70~79',
                'date_start':(today - datetime.timedelta(days=365.25 * 70 ) ).date(),
                'date_end': (today - datetime.timedelta(days=365.25 * 80 ) ).date(),
            },
            {#80세 이상
                'name':'80~',
                'date_start':(today - datetime.timedelta(days=365.25 * 80 ) ).date(),
                'date_end': (today - datetime.timedelta(days=365.25 * 150 ) ).date(),
            },
        ]



    for age in age_list: 
        age_count = Reception.objects.filter(
                **kwargs ,
                recorded_date__range = (date_min, date_max), 
                patient__date_of_birth__lte = age['date_start'].strftime("%Y-%m-%d"),
                patient__date_of_birth__gt = age['date_end'].strftime("%Y-%m-%d"),
            ).exclude(
                progress='deleted'
            ).select_related(
                'patient'
            ).prefetch_related(
                'payment__paymentrecord_set'
            ).aggregate(
                count = Count('id'),
                total_price = Sum('payment__paymentrecord__paid')
            )

        data_list_age.append({
            'name':age['name'],
            'count':age_count['count'],
            'price_sum':0 if age_count['total_price'] is None else age_count['total_price'],
            })


    return JsonResponse({
        'result':True,
        'datas_gender':data_list_gender,
        'datas_nation':data_list_nation,
        'datas_payment_method':data_list_payment_method,
        'datas_age':data_list_age,

        })



def statistics_ymw(request):
    
    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })

    
    
    return render(request, 
    'statistics/statistics_ymw.html',
        {
            'depart_medical':depart_medical,
        }
    )

def search_ymw(request):


    year = request.POST.get('year')
    depart = request.POST.get('depart')

    kwargs = {}
    if depart!='':
        kwargs['depart'] = depart # 기본 


    start = datetime.date(year = int(year), month=1,day=1)
    end = datetime.date(year = int(year), month=12,day=31)

    date_min = datetime.datetime.combine(start, datetime.time.min)
    date_max = datetime.datetime.combine(end, datetime.time.max)

    print(date_min)
    print(date_max)


    data_list_year = []
    data_list_monthly = []
    data_list_week = []
    data_list_hour = []

    query = Reception.objects.filter(
            **kwargs ,
            recorded_date__range = (date_min, date_max), 
        ).exclude(
            progress='deleted'
        ).prefetch_related(
            'payment__paymentrecord_set'
        )

    #년도별
    year_query = query.aggregate(
                count = Count('id'),
                total_price = Sum('payment__paymentrecord__paid')
            )

    data_list_year.append({
            'name':year ,
            'count':year_query['count'],
            'price_sum':0 if year_query['total_price'] is None else year_query['total_price'],
        })


    #월 별
    list_month = [
        {
            'name':'1',
            'date_start':datetime.datetime(year=int(year), month=1,day=1),
            'date_end': datetime.datetime(year=int(year), month=2,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'2',
            'date_start':datetime.datetime(year=int(year), month=2,day=1),
            'date_end': datetime.datetime(year=int(year), month=3,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'3',
            'date_start':datetime.datetime(year=int(year), month=3,day=1),
            'date_end': datetime.datetime(year=int(year), month=4,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'4',
            'date_start':datetime.datetime(year=int(year), month=4,day=1),
            'date_end': datetime.datetime(year=int(year), month=5,day=1) - datetime.timedelta(seconds = 1),
        },       
        {
            'name':'5',
            'date_start':datetime.datetime(year=int(year), month=5,day=1),
            'date_end': datetime.datetime(year=int(year), month=6,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'6',
            'date_start':datetime.datetime(year=int(year), month=6,day=1),
            'date_end': datetime.datetime(year=int(year), month=7,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'7',
            'date_start':datetime.datetime(year=int(year), month=7,day=1),
            'date_end': datetime.datetime(year=int(year), month=8,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'8',
            'date_start':datetime.datetime(year=int(year), month=8,day=1),
            'date_end': datetime.datetime(year=int(year), month=9,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'9',
            'date_start':datetime.datetime(year=int(year), month=9,day=1),
            'date_end': datetime.datetime(year=int(year), month=10,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'10',
            'date_start':datetime.datetime(year=int(year), month=10,day=1),
            'date_end': datetime.datetime(year=int(year), month=11,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'11',
            'date_start':datetime.datetime(year=int(year), month=11,day=1),
            'date_end': datetime.datetime(year=int(year), month=12,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'12',
            'date_start':datetime.datetime(year=int(year), month=12,day=1),
            'date_end': datetime.datetime(year=int(year) + 1, month=1,day=1) - datetime.timedelta(seconds = 1),
        },
    ]


    
    for month in list_month: 
        year_query = query.filter(
            recorded_date__gte = month['date_start'],
            recorded_date__lt = month['date_end'],
            ).aggregate(
                count = Count('id'),
                total_price = Sum('payment__paymentrecord__paid')
            )

        data_list_monthly.append({
            'name':month['name'],
            'count':year_query['count'],
            'price_sum':0 if year_query['total_price'] is None else year_query['total_price'],

            })







    #요일별: 
    list_week= [
        {
            'name':_('Monday'),
            'value':2
        },
        {
            'name':_('Tuesday'),
            'value':3,
        },
        {
            'name':_('Wednesday'),
            'value':4,
        },
        {
            'name':_('Thursday'),
            'value':5,
        },       
        {
            'name':_('Friday'),
            'value':6,
        },
        {
            'name':_('Saturday'),
            'value':7,
        },
        {
            'name':_('Sunday'),
            'value':1,
        },
    ]

    for data in list_week:
        week_query = query.filter(
            recorded_date__week_day = data['value'] ,
            ).aggregate(
                count = Count('id'),
                total_price = Sum('payment__paymentrecord__paid')
            )

        data_list_week.append({
                'name':data['name'],
                'count':week_query['count'],
                'price_sum':0 if week_query['total_price'] is None else week_query['total_price'],
            })



    #시간별: 
    list_hour= [
        {'name':'8','value':8,},
        {'name':'9','value':9,},
        {'name':'10','value':10,},
        {'name':'11','value':11,},
        {'name':'12','value':12,},
        {'name':'13','value':13,},
        {'name':'14','value':14,},
        {'name':'15','value':15,},
        {'name':'16','value':16,},
        {'name':'17','value':17,},
        {'name':'18','value':18,},
        {'name':'19','value':19,},
        
    ]

    
    for data in list_hour:
        hour_query = query.filter(
            recorded_date__hour = data['value'] ,
            ).aggregate(
                count = Count('id'),
                total_price = Sum('payment__paymentrecord__paid')
            )

        data_list_hour.append({
                'name':data['name'],
                'count':hour_query['count'],
                'price_sum':0 if hour_query['total_price'] is None else hour_query['total_price'],
            })



    return JsonResponse({
        'result':True,
        'datas_year':data_list_year,
        'datas_monthly':data_list_monthly,
        'datas_week':data_list_week,
        'datas_hour':data_list_hour,

        })




def statistics_daily(request):


    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })

    today_year = datetime.datetime.now().year
    today_month= datetime.datetime.now().month

    return render(request, 
    'statistics/statistics_daily.html',
        {
            'depart_medical':depart_medical,
            'today_year':today_year,
            'today_month':today_month,
        }
    )

def search_daily(request):

    year = int( request.POST.get('year') )
    month = int(request.POST.get('month') )
    depart = request.POST.get('depart')

    kwargs = {}
    if depart!='':
        kwargs['depart'] = depart # 기본 


    date_count = calendar.monthrange(int(year), int(month))[1]



    #date_min = datetime.datetime.combine(datetime.datetime.strptime(start,"%Y-%m-%d").date(), datetime.time.min)
    #date_max = datetime.datetime.combine(datetime.datetime.strptime(end,"%Y-%m-%d").date(), datetime.time.max)


    query = Reception.objects.filter(
            **kwargs ,
            recorded_date__year = year,
            recorded_date__month= month,
        ).exclude(
            progress='deleted'
        ).prefetch_related(
            'payment__paymentrecord_set'
        )


    list_date_count = []
    list_date_revenue = []
    for count in range(1,date_count + 1):
        tmp_date = datetime.datetime(year = year, month = month, day = count)

        date_query = query.filter(
            recorded_date__date = tmp_date,
            ).aggregate(
                count = Count('id'),
                total_price = Sum('payment__paymentrecord__paid')
            )

        if date_query['count'] is 0 and date_query['total_price'] is None:
            continue


        list_date_count.append({
            'id':count,
            'start':tmp_date.strftime('%Y-%m-%d'),
            'title':'Visits : ' + str(date_query['count'])
            })

        list_date_count.append({
            'id':count,
            'start':tmp_date.strftime('%Y-%m-%d'),
            'title':'Amount : ' + str(0 if date_query['total_price'] is None else "{:,}".format(date_query['total_price'])) + " VND",
            'backgroundColor':'rgb(254,154,202)',
            'borderColor':'rgb(254,154,202)',
            })


    return JsonResponse({
        'result':True,
        'datas_date_count':list_date_count,
        'datas_date_revenue':list_date_revenue,
        })



@login_required
def code_setting(request):

    return render(request,
    'Manage/code_setting.html',
            {
                '':None,
            },
        )

@login_required
def code_search(request):


    
    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')

    
    string=request.POST.get("string")

    kwargs = {}
    argument_list = [] 
    #if string != '':
    #argument_list.append( Q(**{'name_kor__icontains':string} ) )
    argument_list.append( Q(**{'commcode_name_ko__icontains':string} ) )
    argument_list.append( Q(**{'commcode_name_en__icontains':string} ) )
    argument_list.append( Q(**{'commcode_name_vi__icontains':string} ) )

    #if project_type != '':
    #    kwargs['type'] = project_type

    datas = []
    query = COMMCODE.objects.filter(
        functools.reduce(operator.or_, argument_list),
        **kwargs,
        use_yn = 'Y')







    for data in query:

        datas.append({
            'id':data.id,

            'upper_commcode':data.upper_commcode,
            'upper_commcode_name':data.upper_commcode_name,
            'commcode_grp':data.commcode_grp,
            'commcode_grp_name':data.commcode_grp_name,
            'commcode':data.commcode,
            'commcode_name_ko':data.commcode_name_ko,
            'commcode_name_en':data.commcode_name_en,
            'commcode_name_vi':data.commcode_name_vi,
            'se1':data.se1,
            'se2':data.se2,
            'se3':data.se3,
            'se4':data.se4,
            'se5':data.se5,
            'se6':data.se6,
            'se7':data.se7,
            'se8':data.se8,
            'seq':data.seq,
            'registrerer':data.registrerer,
            'date_of_registered':data.date_of_registered.strftime('%Y-%m-%d'),
            })


    page_context = request.POST.get('page_context',10) # 페이지 컨텐츠 
    page = request.POST.get('page',1)

    paginator = Paginator(datas, page_context)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)


    context = {
            'datas':list(paging_data),
            'page_range_start':paging_data.paginator.page_range.start,
            'page_range_stop':paging_data.paginator.page_range.stop,
            'page_number':paging_data.number,
            'has_previous':paging_data.has_previous(),
            'has_next':paging_data.has_next(),



            }



    return JsonResponse(context)


@login_required
def code_save(request):

    
    id=request.POST.get("id")

    code_upper_commcode=request.POST.get("code_upper_commcode")
    code_upper_commcode_name=request.POST.get("code_upper_commcode_name")
    code_commcode_grp=request.POST.get("code_commcode_grp")
    code_commcode_grp_name=request.POST.get("code_commcode_grp_name")
    code_commcode=request.POST.get("code_commcode")
    code_commcode_name_ko=request.POST.get("code_commcode_name_ko")
    code_commcode_name_en=request.POST.get("code_commcode_name_en")
    code_commcode_name_vi=request.POST.get("code_commcode_name_vi")
    code_se1=request.POST.get("code_se1")
    code_se2=request.POST.get("code_se2")
    code_se3=request.POST.get("code_se3")
    code_se4=request.POST.get("code_se4")
    code_se5=request.POST.get("code_se5")
    code_se6=request.POST.get("code_se6")
    code_se7=request.POST.get("code_se7")
    code_se8=request.POST.get("code_se8")
    code_seq=request.POST.get("code_seq")
    code_use_yn=request.POST.get("code_use_yn")


    if id == '':
        commcode = COMMCODE()

        commcode.registrerer = request.user.id
    else:
        commcode = COMMCODE.objects.get(id = id)


    commcode.upper_commcode = code_upper_commcode
    commcode.upper_commcode_name = code_upper_commcode_name
    commcode.commcode_grp = code_commcode_grp
    commcode.commcode_grp_name = code_commcode_grp_name
    commcode.commcode = code_commcode
    commcode.commcode_name_ko = code_commcode_name_ko
    commcode.commcode_name_en = code_commcode_name_en
    commcode.commcode_name_vi = code_commcode_name_vi
    commcode.se1 = code_se1
    commcode.se2 = code_se2
    commcode.se3 = code_se3
    commcode.se4 = code_se4
    commcode.se5 = code_se5
    commcode.se6 = code_se6
    commcode.se7 = code_se7
    commcode.se8 = code_se8
    commcode.seq = code_seq

    commcode.use_yn = code_use_yn

    commcode.lastest_modifier = request.user.id
    commcode.lastest_modified_date = datetime.datetime.now()

    commcode.save()



    return JsonResponse({
        'result':True,        
        })


@login_required
def code_get(request):

    
    id=request.POST.get("id")

    code = COMMCODE.objects.get(id = id)



    return JsonResponse({

        "code_upper_commcode":code.upper_commcode,
        "code_upper_commcode_name":code.upper_commcode_name,
        "code_commcode_grp":code.commcode_grp,
        "code_commcode_grp_name":code.commcode_grp_name,
        "code_commcode":code.commcode,
        "code_commcode_name_ko":code.commcode_name_ko,
        "code_commcode_name_en":code.commcode_name_en,
        "code_commcode_name_vi":code.commcode_name_vi,
        "code_se1":code.se1,
        "code_se2":code.se2,
        "code_se3":code.se3,
        "code_se4":code.se4,
        "code_se5":code.se5,
        "code_se6":code.se6,
        "code_se7":code.se7,
        "code_se8":code.se8,
        "code_seq":code.seq,
        "code_use_yn":code.use_yn,



        })




@login_required
def code_delete(request):

    id=request.POST.get("id")

    code = COMMCODE.objects.get(id = id)
    code.use_yn = 'N'

    code.lastest_modifier = request.user.id
    code.lastest_modified_date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    code.save()


    return JsonResponse({
        'result':True,    
        })
